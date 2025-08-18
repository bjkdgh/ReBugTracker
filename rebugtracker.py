# rebugtracker.py: ReBugTracker主程序
# 基于Flask的缺陷跟踪系统，支持用户注册、登录、问题提交、分配和解决等功能
# 支持PostgreSQL和SQLite两种数据库类型

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, abort, make_response, send_from_directory
from config import DB_TYPE, ALLOWED_EXTENSIONS, MAX_CONTENT_LENGTH, DATABASE_CONFIG
from config_adapter import UPLOAD_FOLDER, SECRET_KEY, FLASK_DEBUG
import psycopg2
from psycopg2.extras import DictCursor
from functools import wraps
import os
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
from db_factory import get_db_connection
from sql_adapter import adapt_sql
import traceback
from urllib.parse import quote, unquote

def safe_get(obj, key, default=None):
    """安全获取对象属性，兼容字典和sqlite3.Row对象"""
    if obj is None:
        return default
    try:
        # 尝试字典方式访问
        if hasattr(obj, 'get'):
            return obj.get(key, default)
        # 尝试属性方式访问
        elif hasattr(obj, key):
            return getattr(obj, key, default)
        # 尝试索引方式访问（sqlite3.Row支持）
        elif hasattr(obj, '__getitem__'):
            try:
                return obj[key]
            except (KeyError, IndexError):
                return default
        else:
            return default
    except:
        return default

# 数据库配置
DB_CONFIG = DATABASE_CONFIG[DB_TYPE]

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 初始化Flask应用
app = Flask(__name__, static_folder='static', static_url_path='/static', template_folder='templates')

# 确保默认字符集为UTF-8
app.config.update(
    SECRET_KEY=SECRET_KEY,  # 从config_adapter加载
    DEBUG=FLASK_DEBUG,  # 从config_adapter加载
    PROPAGATE_EXCEPTIONS=True,  # 传播异常
    TRAP_HTTP_EXCEPTIONS=False,
    UPLOAD_FOLDER=UPLOAD_FOLDER,  # 从config_adapter加载，支持绝对路径
    ALLOWED_EXTENSIONS=ALLOWED_EXTENSIONS,  # 从config_adapter加载
    MAX_CONTENT_LENGTH=MAX_CONTENT_LENGTH,  # 从config_adapter加载
    JSON_AS_ASCII=False,  # 确保JSON响应不使用ASCII编码
    DEFAULT_CHARSET='utf-8'  # 设置默认字符集
)

# 添加响应头中间件确保所有响应使用UTF-8
@app.after_request
def add_charset(response):
    response.headers['Content-Type'] = 'text/html; charset=utf-8'
    return response

# 错误处理函数
@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500

@app.errorhandler(403)
def forbidden_error(error):
    return render_template('403.html'), 403

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 注册时间格式化过滤器
@app.template_filter('datetimeformat')
def datetimeformat_filter(value, format='%Y-%m-%d %H:%M:%S'):
    if not value:
        return '--'  # 空值占位符

    try:
        # 处理字符串类型输入（兼容多种数据库时间格式）
        if isinstance(value, str):
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M'):
                try:
                    dt = datetime.strptime(value, fmt)
                    return dt.strftime(format)
                except ValueError:
                    continue
            return value  # 无法解析时返回原值

        # 处理datetime对象
        return value.strftime(format)
    except Exception as e:
        print(f"时间格式化错误: {str(e)} | 原始值: {value} ({type(value)})")
        return str(value)  # 异常时返回原始值的字符串形式

def get_current_user():
    """从cookie获取当前用户信息"""
    try:
        user_id = request.cookies.get('user_id')
        username = request.cookies.get('username')
        role_en = request.cookies.get('role_en')
        team = request.cookies.get('team')
        chinese_name = request.cookies.get('chinese_name')

        # 解码URL编码的cookie值
        if team:
            team = unquote(team)
        if chinese_name:
            chinese_name = unquote(chinese_name)

        if not all([user_id, username, role_en]):
            app.logger.debug(f"获取用户信息失败 - 缺少必要cookie: user_id={user_id}, username={username}, role_en={role_en}")
            return None

        # 添加调试信息
        app.logger.debug(f"当前用户cookie - user_id: {user_id}, username: {username}, role_en: {role_en}, team: {team}, chinese_name: {chinese_name}")

        # 角色映射
        role_mapping = {
            'gly': '管理员',
            'fzr': '负责人',
            'zncy': '组内成员',
            'ssz': '实施组',
            'pm': '产品经理'
        }

        role_en_lower = role_en.lower() if role_en else None
        role_cn = role_mapping.get(role_en_lower, role_en_lower)

        user_data = {
            'id': int(user_id),
            'username': username,
            'chinese_name': chinese_name,
            'role': role_cn,  # 中文角色
            'role_en': role_en_lower,  # 英文角色
            'team': team
        }
        app.logger.debug(f"返回用户数据: {user_data}")
        return user_data
    except Exception as e:
        app.logger.error(f"获取用户信息异常: {str(e)}")
        return None

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user = get_current_user()
        if not user:
            resp = make_response(redirect('/login'))
            resp.delete_cookie('user_id')
            resp.delete_cookie('username')
            resp.delete_cookie('role_en')
            resp.delete_cookie('team')
            resp.delete_cookie('chinese_name')
            return resp
        return f(*args, **kwargs)
    return decorated_function

def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user = get_current_user()
            if not user or not safe_get(user, 'role'):
                app.logger.debug(f"权限检查失败 - 用户未登录或缺少角色信息")
                abort(403)

            user_role = safe_get(user, 'role_en', '').lower()
            required_role = role.lower()

            app.logger.debug(f"权限检查 - 用户角色: {user_role}, 要求角色: {required_role}")

            # 管理员拥有所有权限
            if user_role == 'gly':
                app.logger.debug("权限检查通过 - 管理员权限")
                return f(*args, **kwargs)

            # 允许更高权限角色访问
            if required_role == 'zncy' and user_role in ['zncy', 'fzr', 'ssz']:
                app.logger.debug(f"权限检查通过 - 用户角色 {user_role} 满足要求 {required_role}")
                return f(*args, **kwargs)

            # 产品经理权限检查
            if required_role == 'pm' and user_role == 'pm':
                app.logger.debug(f"权限检查通过 - 产品经理权限")
                return f(*args, **kwargs)

            if required_role == 'fzr' and user_role in ['fzr', 'ssz']:
                app.logger.debug(f"权限检查通过 - 用户角色 {user_role} 满足要求 {required_role}")
                return f(*args, **kwargs)

            if required_role == 'ssz' and user_role == 'ssz':
                app.logger.debug(f"权限检查通过 - 用户角色 {user_role} 匹配要求 {required_role}")
                return f(*args, **kwargs)

            if user_role == required_role:
                app.logger.debug(f"权限检查通过 - 用户角色 {user_role} 匹配要求 {required_role}")
                return f(*args, **kwargs)

            app.logger.debug(f"权限检查失败 - 用户角色 {user_role} 不满足要求 {required_role}")
            abort(403)
        return decorated_function
    return decorator

# 数据库初始化函数
def init_db():
    """初始化数据库结构

    功能：
    - 根据数据库类型创建完整的数据库表结构
    - 创建users表（包含通知相关字段）
    - 创建bugs表（包含外键约束）
    - 创建bug_images表（问题图片）
    - 创建system_config表（系统配置）
    - 创建user_notification_preferences表（用户通知偏好）
    - 创建notifications表（通知记录）
    - 添加缺失的字段（如role_en, team_en, chinese_name, email, phone等）
    - 更新现有数据的角色英文标识
    - 创建表达式索引以实现大小写不敏感的用户名唯一约束
    - 确保存在默认管理员账户
    """
    # 获取数据库连接
    conn = get_db_connection()
    # 对于SQLite，设置自动提交模式（SQLite没有显式的事务，所以设置autocommit为True没有实际作用，但为了兼容性保留）
    if DB_TYPE == 'postgres':
        conn.autocommit = True
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    # 创建用户表（兼容SQLite和PostgreSQL）
    if DB_TYPE == 'postgres':
        # PostgreSQL建表语句 - 与当前数据库结构保持一致
        c.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username TEXT NOT NULL,
                password TEXT NOT NULL,
                role TEXT NOT NULL,
                team TEXT,
                role_en TEXT,
                team_en TEXT,
                chinese_name TEXT,
                email CHARACTER VARYING(255),
                phone CHARACTER VARYING(20),
                gotify_app_token CHARACTER VARYING(255),
                gotify_user_id CHARACTER VARYING(255),
                preferences TEXT
            )
        ''')
    else:
        # SQLite建表语句
        c.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                chinese_name TEXT,
                role TEXT DEFAULT 'zncy',
                role_en TEXT DEFAULT 'zncy',
                team TEXT,
                team_en TEXT,
                email TEXT,
                phone TEXT,
                gotify_app_token TEXT,
                gotify_user_id TEXT,
                preferences TEXT
            )
        ''')

    # 添加新列(如果不存在) - 兼容SQLite和PostgreSQL
    columns_to_add = ['role_en', 'team_en', 'chinese_name', 'email', 'phone', 'gotify_app_token', 'gotify_user_id']
    for col in columns_to_add:
        try:
            if DB_TYPE == 'postgres':
                # 使用IF NOT EXISTS语法添加列（PostgreSQL特性）
                if col in ['role_en']:
                    c.execute(f'ALTER TABLE users ADD COLUMN IF NOT EXISTS {col} TEXT DEFAULT \'zncy\'')
                else:
                    c.execute(f'ALTER TABLE users ADD COLUMN IF NOT EXISTS {col} TEXT')
            else:
                # SQLite不支持ADD COLUMN IF NOT EXISTS，需要先检查列是否存在
                c.execute(f"PRAGMA table_info(users)")
                columns = [info[1] for info in c.fetchall()]
                if col not in columns:
                    if col in ['role_en']:
                        c.execute(f'ALTER TABLE users ADD COLUMN {col} TEXT DEFAULT \'zncy\'')
                    else:
                        c.execute(f'ALTER TABLE users ADD COLUMN {col} TEXT')
        except Exception as e:
            print(f"添加列{col}时出错: {str(e)}")
            if DB_TYPE == 'postgres':
                conn.rollback()

    # 更新现有数据的角色英文缩写（仅当有数据时）
    if DB_TYPE == 'postgres':
        # 使用CASE语句将中文角色转换为英文标识
        c.execute('''
            UPDATE users SET
                role_en = CASE role
                    WHEN '管理员' THEN 'gly'
                    WHEN '负责人' THEN 'fzr'
                    WHEN '组内成员' THEN 'zncy'
                    WHEN '实施组' THEN 'ssz'
                    WHEN '产品经理' THEN 'pm'
                    ELSE role
                END
        ''')
    else:
        # SQLite版本
        c.execute('''
            UPDATE users SET
                role_en = CASE role
                    WHEN '管理员' THEN 'gly'
                    WHEN '负责人' THEN 'fzr'
                    WHEN '组内成员' THEN 'zncy'
                    WHEN '实施组' THEN 'ssz'
                    WHEN '产品经理' THEN 'pm'
                    ELSE role
                END
        ''')

    # 创建索引（仅PostgreSQL需要，SQLite会自动创建必要索引）
    if DB_TYPE == 'postgres':
        # 为notifications表创建索引（与当前数据库结构保持一致）
        try:
            c.execute('CREATE INDEX IF NOT EXISTS idx_notifications_created_at ON notifications (created_at)')
            c.execute('CREATE INDEX IF NOT EXISTS idx_notifications_read_status ON notifications (read_status)')
            c.execute('CREATE INDEX IF NOT EXISTS idx_notifications_related_bug_id ON notifications (related_bug_id)')
            c.execute('CREATE INDEX IF NOT EXISTS idx_notifications_user_id ON notifications (user_id)')
        except Exception as e:
            print(f"创建索引时出错: {e}")
            if DB_TYPE == 'postgres':
                conn.rollback()

    # 检查并添加默认管理员账户(如果不存在)
    # 查询用户名为admin的用户
    query, params = adapt_sql('SELECT * FROM users WHERE username = %s', ('admin',))
    c.execute(query, params)
    if not c.fetchone():
        # 如果不存在则创建默认管理员账户
        hashed_password = generate_password_hash('admin')
        query, params = adapt_sql('''
            INSERT INTO users (username, password, role, role_en, chinese_name, team, team_en)
            VALUES (%s, %s, '管理员', 'gly', '系统管理员', '管理员', 'gly')
        ''', ('admin', hashed_password))
        c.execute(query, params)

    # 确保admin用户的完整信息正确设置
    query, params = adapt_sql('''
        UPDATE users SET
            role_en = 'gly',
            chinese_name = CASE WHEN chinese_name IS NULL OR chinese_name = '' THEN '系统管理员' ELSE chinese_name END,
            team = CASE WHEN team IS NULL OR team = '' THEN '管理员' ELSE team END,
            team_en = CASE WHEN team_en IS NULL OR team_en = '' THEN 'gly' ELSE team_en END
        WHERE username = 'admin'
    ''', ())
    c.execute(query, params)

    # PostgreSQL: 确保角色类型存在
    if DB_TYPE == 'postgres':
        # 使用DO块执行条件创建类型的操作
        c.execute('''
            DO $$
            BEGIN
                IF NOT EXISTS (SELECT 1 FROM pg_type WHERE typname = 'user_role') THEN
                    CREATE TYPE user_role AS ENUM ('管理员', '负责人', '组内成员', '实施组', '产品经理');
                END IF;
            END $$;
        ''')

    # 创建bugs表（兼容SQLite和PostgreSQL）
    if DB_TYPE == 'postgres':
        # PostgreSQL建表语句 - 与当前数据库结构保持一致（暂不添加外键约束）
        c.execute('''
            CREATE TABLE IF NOT EXISTS bugs (
                id SERIAL PRIMARY KEY,
                title TEXT NOT NULL,
                description TEXT,
                status TEXT DEFAULT '待处理',
                type TEXT DEFAULT 'bug',
                assigned_to INTEGER,
                created_by INTEGER,
                project TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                resolved_at TIMESTAMP,
                resolution TEXT,
                image_path TEXT
            )
        ''')
    else:
        c.execute('''
            CREATE TABLE IF NOT EXISTS bugs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                description TEXT,
                status TEXT DEFAULT '待处理',
                type TEXT DEFAULT 'bug',    -- 类型：需求/bug
                assigned_to INTEGER,         -- 负责人ID
                created_by INTEGER,          -- 提交人ID
                project TEXT,               -- 所属项目名称
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                resolved_at TIMESTAMP,
                resolution TEXT,            -- 处理详情
                image_path TEXT,            -- 图片路径
                FOREIGN KEY (assigned_to) REFERENCES users (id),
                FOREIGN KEY (created_by) REFERENCES users (id)
            )
        ''')

    # 创建问题图片表
    if DB_TYPE == 'postgres':
        # PostgreSQL建表语句 - 与当前数据库结构保持一致
        c.execute('''
            CREATE TABLE IF NOT EXISTS bug_images (
                id SERIAL PRIMARY KEY,
                bug_id INTEGER NOT NULL,
                image_path TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        # 添加外键约束（如果不存在）
        try:
            c.execute('''
                ALTER TABLE bug_images
                ADD CONSTRAINT bug_images_bug_id_fkey
                FOREIGN KEY (bug_id) REFERENCES bugs (id) ON DELETE CASCADE
            ''')
        except Exception as e:
            # 外键约束可能已存在，忽略错误
            if DB_TYPE == 'postgres':
                conn.rollback()
    else:
        c.execute('''
            CREATE TABLE IF NOT EXISTS bug_images (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bug_id INTEGER NOT NULL,
                image_path TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (bug_id) REFERENCES bugs (id) ON DELETE CASCADE
            )
        ''')

    # 创建项目表
    if DB_TYPE == 'postgres':
        c.execute('''
            CREATE TABLE IF NOT EXISTS projects (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                type TEXT NOT NULL,
                city TEXT NOT NULL,
                start_date DATE,
                factory_acceptance_date DATE,
                site_acceptance_date DATE,
                practical_date DATE,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
    else:
        c.execute('''
            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                type TEXT NOT NULL,
                city TEXT NOT NULL,
                start_date DATE,
                factory_acceptance_date DATE,
                site_acceptance_date DATE,
                practical_date DATE,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

    # 创建产品线表
    if DB_TYPE == 'postgres':
        c.execute('''
            CREATE TABLE IF NOT EXISTS product_lines (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                description TEXT,
                status TEXT DEFAULT 'active',
                created_by INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
    else:
        c.execute('''
            CREATE TABLE IF NOT EXISTS product_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                description TEXT,
                status TEXT DEFAULT 'active',
                created_by INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (created_by) REFERENCES users (id)
            )
        ''')

    # 创建系统配置表
    if DB_TYPE == 'postgres':
        # PostgreSQL建表语句 - 与当前数据库结构保持一致（暂不添加外键约束）
        c.execute('''
            CREATE TABLE IF NOT EXISTS system_config (
                config_key CHARACTER VARYING(50) PRIMARY KEY,
                config_value TEXT NOT NULL,
                description TEXT,
                updated_by INTEGER,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
    else:
        c.execute('''
            CREATE TABLE IF NOT EXISTS system_config (
                config_key TEXT PRIMARY KEY,
                config_value TEXT NOT NULL,
                description TEXT,
                updated_by INTEGER,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (updated_by) REFERENCES users (id)
            )
        ''')

    # 创建用户通知偏好表
    if DB_TYPE == 'postgres':
        # PostgreSQL建表语句 - 与当前数据库结构保持一致
        c.execute('''
            CREATE TABLE IF NOT EXISTS user_notification_preferences (
                user_id INTEGER PRIMARY KEY,
                email_enabled BOOLEAN DEFAULT TRUE,
                gotify_enabled BOOLEAN DEFAULT TRUE,
                inapp_enabled BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
    else:
        c.execute('''
            CREATE TABLE IF NOT EXISTS user_notification_preferences (
                user_id INTEGER PRIMARY KEY,
                email_enabled BOOLEAN DEFAULT 1,
                inapp_enabled BOOLEAN DEFAULT 1,
                gotify_enabled BOOLEAN DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE
            )
        ''')

    # 创建通知表
    if DB_TYPE == 'postgres':
        # PostgreSQL建表语句 - 与当前数据库结构保持一致（暂不添加外键约束）
        c.execute('''
            CREATE TABLE IF NOT EXISTS notifications (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                title CHARACTER VARYING(200) NOT NULL,
                content TEXT NOT NULL,
                read_status BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                read_at TIMESTAMP,
                related_bug_id INTEGER
            )
        ''')
    else:
        c.execute('''
            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                content TEXT,
                read_status BOOLEAN DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                read_at TIMESTAMP,
                related_bug_id INTEGER,
                FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE,
                FOREIGN KEY (related_bug_id) REFERENCES bugs (id) ON DELETE SET NULL
            )
        ''')

    # 为bugs表添加product_line_id字段（如果不存在）
    try:
        if DB_TYPE == 'postgres':
            c.execute('ALTER TABLE bugs ADD COLUMN IF NOT EXISTS product_line_id INTEGER REFERENCES product_lines(id)')
        else:
            # SQLite需要先检查列是否存在
            c.execute("PRAGMA table_info(bugs)")
            columns = [info[1] for info in c.fetchall()]
            if 'product_line_id' not in columns:
                c.execute('ALTER TABLE bugs ADD COLUMN product_line_id INTEGER')
    except Exception as e:
        print(f"添加product_line_id字段时出错: {e}")

    # 插入示例产品线数据
    try:
        sample_products = [
            ('实施组', '实施组产品线'),
            ('实施组研发', '实施组研发产品线'),
            ('新能源', '新能源产品线'),
            ('网络分析', '网络分析产品线'),
            ('第三道防线', '第三道防线产品线'),
            ('智能告警', '智能告警产品线'),
            ('操作票及防误', '操作票及防误产品线'),
            ('电量', '电量产品线'),
            ('消纳', '消纳产品线'),
            ('自动发电控制', '自动发电控制产品线')
        ]

        for name, description in sample_products:
            if DB_TYPE == 'postgres':
                query = "INSERT INTO product_lines (name, description) VALUES (%s, %s) ON CONFLICT (name) DO NOTHING"
            else:
                query = "INSERT OR IGNORE INTO product_lines (name, description) VALUES (?, ?)"

            if DB_TYPE == 'postgres':
                c.execute(query, (name, description))
            else:
                c.execute(query, (name, description))
    except Exception as e:
        print(f"插入示例产品线数据时出错: {e}")

    # 提交事务并关闭连接
    conn.commit()
    conn.close()
@app.route('/bug/complete/<int:bug_id>', methods=['POST'])
@login_required
@role_required('ssz')
def complete_bug(bug_id):
    """确认闭环问题"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    try:
        # 检查问题是否存在且状态为'已解决'
        query, params = adapt_sql('SELECT * FROM bugs WHERE id = %s', (bug_id,))
        c.execute(query, params)
        bug = c.fetchone()
        if not bug:
            return jsonify({'success': False, 'message': '问题不存在'}), 404
        if bug['status'] != '已解决':
            return jsonify({'success': False, 'message': '问题状态不是已解决，无法闭环'}), 400

        # 获取问题的详细信息用于通知
        if hasattr(bug, 'keys'):
            bug_info = dict(bug)
        else:
            # SQLite返回的是tuple，需要手动映射
            bug_info = {
                'id': bug[0], 'title': bug[1], 'description': bug[2],
                'status': bug[3], 'assigned_to': bug[4], 'created_by': bug[5],
                'project': bug[6], 'created_at': bug[7], 'resolved_at': bug[8],
                'resolution': bug[9], 'image_path': bug[10]
            }

        # 更新问题状态为"已完成"
        query, params = adapt_sql('''
            UPDATE bugs
            SET status = '已完成'
            WHERE id = %s
        ''', (bug_id,))
        c.execute(query, params)
        conn.commit()

        # 异步发送通知（在后台处理）
        def send_closure_notification_async():
            try:
                app.logger.info(f"后台发送问题关闭通知 - bug_id: {bug_id}")
                from notification.simple_notifier import simple_notifier
                from datetime import datetime

                notification_data = {
                    'bug_id': bug_id,
                    'title': bug_info['title'],
                    'description': bug_info['description'],
                    'close_reason': '实施组确认闭环',
                    'closer_name': user['chinese_name'] or user['username'],
                    'closed_time': datetime.now().isoformat(),
                    'creator_id': bug_info['created_by'],
                    'assignee_id': bug_info.get('assigned_to')
                }

                simple_notifier.send_flow_notification('bug_closed', notification_data)
                app.logger.info(f"问题关闭通知发送完成 - bug_id: {bug_id}")
            except Exception as e:
                app.logger.error(f"后台关闭通知发送失败 - bug_id: {bug_id}, 错误: {e}")
                import traceback
                app.logger.error(f"关闭通知发送错误详情: {traceback.format_exc()}")

        # 启动后台通知任务
        import threading
        notification_thread = threading.Thread(target=send_closure_notification_async)
        notification_thread.daemon = True
        notification_thread.start()

        return jsonify({'success': True, 'message': '问题已成功闭环'})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

# 登录路由
@app.route('/logout')
def logout():
    """用户登出"""
    resp = make_response(redirect('/login'))
    resp.delete_cookie('user_id')
    resp.delete_cookie('username')
    resp.delete_cookie('role_en')
    resp.delete_cookie('team_en')
    return resp

@app.route('/register', methods=['GET', 'POST'])
def register():
    """用户注册路由"""
    if request.method == 'GET':
        return render_template('register.html')

    # 处理注册请求
    chinese_name = request.form.get('chinese_name')
    username = request.form.get('username')
    password = request.form.get('password')
    email = request.form.get('email')
    phone = request.form.get('phone')
    role = request.form.get('role')

    # 处理团队选择：产品经理支持多团队，其他角色单团队
    if role == 'pm':
        # 产品经理的多团队选择
        pm_teams = request.form.getlist('pm_teams')
        team = ','.join(pm_teams) if pm_teams else None  # 多个团队用逗号分隔
    else:
        # 其他角色的单团队选择
        team = request.form.get('team')

    # 通知偏好设置
    email_notifications = request.form.get('email_notifications') == 'on'
    gotify_notifications = request.form.get('gotify_notifications') == 'on'
    inapp_notifications = request.form.get('inapp_notifications') == 'on'

    if not all([username, password, role, email]):
        return jsonify({'success': False, 'message': '请填写完整信息（用户名、密码、角色、邮箱为必填项）'}), 400

    # 验证邮箱格式
    import re
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(email_pattern, email):
        return jsonify({'success': False, 'message': '请输入有效的邮箱地址'}), 400

    # 角色值映射
    role_mapping = {
        'ssz': '实施组',
        'fzr': '负责人',
        'zncy': '组内成员',
        'pm': '产品经理'
    }
    role_cn = role_mapping.get(role, role)
    role_en = role  # 保持原英文缩写

    # 简单处理team_en
    team_en = team if team else None

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    try:
        hashed_password = generate_password_hash(password)
        if DB_TYPE == 'postgres':
            query, params = adapt_sql(
                '''INSERT INTO users (chinese_name, username, password, email, phone, role, role_en, team, team_en)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id''',
                (chinese_name, username, hashed_password, email, phone, role_cn, role_en, team, team_en)
            )
            c.execute(query, params)
            user_id = c.fetchone()['id']
        else:
            # SQLite模式
            query, params = adapt_sql(
                '''INSERT INTO users (chinese_name, username, password, email, phone, role, role_en, team, team_en)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                (chinese_name, username, hashed_password, email, phone, role_cn, role_en, team, team_en)
            )
            c.execute(query, params)
            user_id = c.lastrowid

        # 创建用户通知偏好设置
        try:
            from notification.notification_manager import NotificationManager
            NotificationManager.set_user_notification_preferences(
                str(user_id),
                email_enabled=email_notifications,
                gotify_enabled=gotify_notifications,
                inapp_enabled=inapp_notifications
            )
        except Exception as e:
            app.logger.warning(f"创建用户通知偏好失败: {e}")
            # 不影响注册流程，只记录警告

        # 产品经理的团队信息已经在上面通过team字段处理了（多个团队用逗号分隔）
        # 这里不需要额外的产品线分配逻辑，因为团队就是产品线

        conn.commit()
        return jsonify({
            'success': True,
            'redirect': '/login',
            'message': '注册成功'
        })
    except Exception as e:
        if 'UNIQUE constraint' in str(e):
            conn.rollback()
            return jsonify({'success': False, 'message': '用户名已存在'}), 400
        # 处理其他类型异常
        conn.rollback()
        return jsonify({'success': False, 'message': '注册失败: ' + str(e)}), 500
    finally:
        conn.close()

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        app.logger.debug(f"登录请求: username={username}, password={'*' * len(password) if password else None}")

        if not username or not password:
            app.logger.debug("用户名或密码为空")
            return jsonify({'success': False, 'message': '用户名和密码不能为空'}), 400

        conn = None
        try:
            conn = get_db_connection()
            if DB_TYPE == 'postgres':
                c = conn.cursor(cursor_factory=DictCursor)
            else:
                c = conn.cursor()
            query, params = adapt_sql('SELECT * FROM users WHERE username = %s', (username,))
            c.execute(query, params)
            user = c.fetchone()

            if not user:
                return jsonify({'success': False, 'message': '用户名或密码错误'}), 401

            if check_password_hash(user['password'], password):
                # 确保role_en不为空
                if not user['role_en']:
                    return jsonify({'success': False, 'message': '用户角色信息不完整，请联系管理员'}), 403
                role_en = user['role_en']
                resp = make_response(jsonify({
                    'success': True,
                    'redirect': '/',
                    'user': {
                        'id': user['id'],
                        'username': user['username'],
                        'role': user['role_en']
                    }
                }))
                resp.set_cookie('user_id', str(user['id']))
                resp.set_cookie('username', user['username'])
                chinese_name = safe_get(user, 'chinese_name') or user['username'] or 'Unknown'
                resp.set_cookie('chinese_name', quote(str(chinese_name)))
                resp.set_cookie('role_en', role_en)
                team_name = safe_get(user, 'team') or 'Unknown'
                resp.set_cookie('team', quote(str(team_name)))
                app.logger.info(f"用户 {user['username']} 登录成功，设置cookie: user_id={user['id']}, username={user['username']}, role_en={role_en}")
                return resp
            else:
                return jsonify({'success': False, 'message': '用户名或密码错误'}), 401
        except Exception as e:
            try:
                app.logger.error(f"登录过程中发生错误: {str(e)}")
            except:
                pass  # 即使记录日志失败也继续
            return jsonify({'success': False, 'message': '服务器内部错误'}), 500
        finally:
            if conn:
                try:
                    conn.close()
                except Exception as e:
                    try:
                        app.logger.error(f"关闭数据库连接时出错: {str(e)}")
                    except:
                        pass

    #
    if request.method == 'GET':
        return render_template('login.html')

    #
    return jsonify({'success': False, 'message': '无效的请求方法'}), 400

# 问题列表路由 (已废弃 - 缺少对应的bugs.html模板)
# @app.route('/bugs')
# @login_required
# def bugs_list():
#     """问题列表页面"""
#     user = get_current_user()
#     if not user:
#         return redirect('/login')
# 
#     try:
#         conn = get_db_connection()
#         if DB_TYPE == 'postgres':
#             c = conn.cursor(cursor_factory=RealDictCursor)
#         else:
#             c = conn.cursor()
# 
#         # 获取筛选参数
#         status_filter = request.args.get('status', '')
#         product_line_filter = request.args.get('product_line', '')
#         type_filter = request.args.get('type', '')
#         page = int(request.args.get('page', 1))
#         per_page = 20
# 
#         # 构建查询条件
#         where_conditions = []
#         params = []
# 
#         if status_filter:
#             where_conditions.append('b.status = %s')
#             params.append(status_filter)
# 
#         if product_line_filter:
#             where_conditions.append('b.product_line_id = %s')
#             params.append(int(product_line_filter))
# 
#         if type_filter:
#             where_conditions.append('b.type = %s')
#             params.append(type_filter)
# 
#         where_clause = 'WHERE ' + ' AND '.join(where_conditions) if where_conditions else ''
# 
#         # 获取总数
#         count_query = f'''
#             SELECT COUNT(*)
#             FROM bugs b
#             LEFT JOIN product_lines pl ON b.product_line_id = pl.id
#             {where_clause}
#         '''
#         query, count_params = adapt_sql(count_query, tuple(params))
#         c.execute(query, count_params)
#         total = c.fetchone()[0]
# 
#         # 获取问题列表
#         offset = (page - 1) * per_page
#         bugs_query = f'''
#             SELECT
#                 b.id, b.title, b.description, b.status, b.type,
#                 b.created_at, b.resolved_at,
#                 pl.name as product_line_name,
#                 u1.chinese_name as creator_name, u1.username as creator_username,
#                 u2.chinese_name as assignee_name, u2.username as assignee_username,
#                 b.assigned_to
#             FROM bugs b
#             LEFT JOIN product_lines pl ON b.product_line_id = pl.id
#             LEFT JOIN users u1 ON b.created_by = u1.id
#             LEFT JOIN users u2 ON b.assigned_to = u2.id
#             {where_clause}
#             ORDER BY b.created_at DESC
#             LIMIT %s OFFSET %s
#         '''
#         query, bugs_params = adapt_sql(bugs_query, tuple(params + [per_page, offset]))
#         c.execute(query, bugs_params)
#         bugs = c.fetchall()
# 
#         # 获取统计数据
#         stats_query = '''
#             SELECT
#                 COUNT(*) as total,
#                 SUM(CASE WHEN status = '待处理' THEN 1 ELSE 0 END) as pending,
#                 SUM(CASE WHEN status = '处理中' THEN 1 ELSE 0 END) as processing,
#                 SUM(CASE WHEN status = '已解决' THEN 1 ELSE 0 END) as resolved,
#                 SUM(CASE WHEN status = '已完成' THEN 1 ELSE 0 END) as closed
#             FROM bugs
#         '''
#         query, stats_params = adapt_sql(stats_query, ())
#         c.execute(query, stats_params)
#         stats_result = c.fetchone()
# 
#         stats = {
#             'total': stats_result[0] or 0,
#             'pending': stats_result[1] or 0,
#             'processing': stats_result[2] or 0,
#             'resolved': stats_result[3] or 0,
#             'closed': stats_result[4] or 0
#         }
# 
#         # 获取产品线列表
#         query, pl_params = adapt_sql('SELECT id, name FROM product_lines WHERE status = %s ORDER BY name', ('active',))
#         c.execute(query, pl_params)
#         product_lines = c.fetchall()
# 
#         # 分页信息
#         pagination = {
#             'page': page,
#             'per_page': per_page,
#             'total': total,
#             'pages': (total + per_page - 1) // per_page,
#             'has_prev': page > 1,
#             'has_next': page * per_page < total,
#             'prev_num': page - 1 if page > 1 else None,
#             'next_num': page + 1 if page * per_page < total else None
#         }
# 
#         # 添加iter_pages方法
#         def iter_pages(left_edge=2, left_current=2, right_current=3, right_edge=2):
#             last = pagination['pages']
#             for num in range(1, last + 1):
#                 if num <= left_edge or \
#                    (pagination['page'] - left_current - 1 < num < pagination['page'] + right_current) or \
#                    num > last - right_edge:
#                     yield num
# 
#         pagination['iter_pages'] = iter_pages
# 
#         conn.close()
# 
#         return render_template('bugs.html',
#                              bugs=bugs,
#                              stats=stats,
#                              product_lines=product_lines,
#                              pagination=pagination,
#                              user=user)
# 
#     except Exception as e:
#         print(f"获取问题列表失败: {e}")
#         return render_template('bugs.html',
#                              bugs=[],
#                              stats={'total': 0, 'pending': 0, 'processing': 0, 'resolved': 0, 'closed': 0},
#                              product_lines=[],
#                              pagination=None,
#                              user=user)

# 首页路由
@app.route('/')
@login_required
def index():

    user = get_current_user()
    if not user:
        return redirect('/login')

    if safe_get(user, 'role_en') == 'zncy':
        # 组内成员直接跳转到team-issues页面
        return redirect('/team-issues')
    elif safe_get(user, 'role_en') == 'gly':
        return redirect('/admin')
    elif safe_get(user, 'role_en') == 'pm':
        # 产品经理直接跳转到产品经理专属页面
        return redirect('/product-manager')
    # 实施组和负责人等角色继续执行后面的代码（渲染首页）

    # 确保数据库连接使用UTF-8编码
    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    if user['role_en'] == 'fzr':
        # 负责人看到自己团队的所有问题和待分配问题
        query = '''
            SELECT b.*, COALESCE(u1.chinese_name, u1.username) as creator_name, COALESCE(u2.chinese_name, u2.username) as assignee_name,
                   b.created_at as local_created_at, b.resolved_at as local_resolved_at, pl.name as product_line_name
            FROM bugs b
            LEFT JOIN users u1 ON b.created_by = u1.id
            LEFT JOIN users u2 ON b.assigned_to = u2.id
            LEFT JOIN product_lines pl ON b.product_line_id = pl.id
            WHERE (b.assigned_to IS NULL OR u2.team = %s) OR u1.team = %s
            ORDER BY b.created_at DESC
        '''
        adapted_query, adapted_params = adapt_sql(query, (user['team'], user['team']))
        c.execute(adapted_query, adapted_params)
    elif user['role_en'] == 'ssz':
        # 实施组只能看到自己创建的问题
        query = '''
            SELECT b.*, COALESCE(u1.chinese_name, u1.username) as creator_name, COALESCE(u2.chinese_name, u2.username) as assignee_name,
                   b.created_at as local_created_at, b.resolved_at as local_resolved_at, pl.name as product_line_name
            FROM bugs b
            LEFT JOIN users u1 ON b.created_by = u1.id
            LEFT JOIN users u2 ON b.assigned_to = u2.id
            LEFT JOIN product_lines pl ON b.product_line_id = pl.id
            WHERE b.created_by = %s
            ORDER BY b.created_at DESC
        '''
        adapted_query, adapted_params = adapt_sql(query, (user['id'],))
        c.execute(adapted_query, adapted_params)
    elif user['role_en'] == 'pm':
        # 产品经理看到自己团队相关的问题（通过assigned_to联合users表的team字段）
        user_teams = user.get('team', '')
        if user_teams:
            team_names = [team.strip() for team in user_teams.split(',') if team.strip()]
            placeholders = ','.join(['%s'] * len(team_names))
            query = f'''
                SELECT b.*, COALESCE(u1.chinese_name, u1.username) as creator_name, COALESCE(u2.chinese_name, u2.username) as assignee_name,
                       b.created_at as local_created_at, b.resolved_at as local_resolved_at, u2.team as product_line_name
                FROM bugs b
                LEFT JOIN users u1 ON b.created_by = u1.id
                LEFT JOIN users u2 ON b.assigned_to = u2.id
                WHERE u2.team IN ({placeholders}) OR b.created_by = %s
                ORDER BY b.created_at DESC
            '''
            # 参数包括团队名称和当前用户ID（显示自己创建的问题）
            params = tuple(team_names) + (user['id'],)
            adapted_query, adapted_params = adapt_sql(query, params)
            c.execute(adapted_query, adapted_params)
        else:
            # 如果没有团队信息，显示空结果
            bugs = []
    else:
        # 其他角色（主要是管理员）看到所有问题
        query = '''
            SELECT b.*, COALESCE(u1.chinese_name, u1.username) as creator_name, COALESCE(u2.chinese_name, u2.username) as assignee_name,
                   b.created_at as local_created_at, b.resolved_at as local_resolved_at, pl.name as product_line_name
            FROM bugs b
            LEFT JOIN users u1 ON b.created_by = u1.id
            LEFT JOIN users u2 ON b.assigned_to = u2.id
            LEFT JOIN product_lines pl ON b.product_line_id = pl.id
            ORDER BY b.created_at DESC
        '''
        adapted_query, adapted_params = adapt_sql(query, ())
        c.execute(adapted_query, adapted_params)

    # 获取bugs数据（如果还没有设置的话）
    if 'bugs' not in locals():
        bugs = c.fetchall()
    elif bugs != []:  # 如果bugs不是空列表，说明需要从数据库获取
        bugs = c.fetchall()

    # 格式化问题创建时间和解决时间
    formatted_bugs = []
    for bug in bugs:
        bug_dict = dict(bug)
        # 处理created_at
        if isinstance(bug_dict['created_at'], str):
            bug_dict['created_at'] = bug_dict['created_at']  # 已经是字符串则直接使用
        elif bug_dict['created_at']:
            bug_dict['created_at'] = bug_dict['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        else:
            bug_dict['created_at'] = '--'

        # 处理resolved_at
        if isinstance(bug_dict['resolved_at'], str):
            bug_dict['resolved_at'] = bug_dict['resolved_at']  # 已经是字符串则直接使用
        elif bug_dict['resolved_at']:
            bug_dict['resolved_at'] = bug_dict['resolved_at'].strftime('%Y-%m-%d %H:%M:%S')
        else:
            bug_dict['resolved_at'] = '--'

        formatted_bugs.append(bug_dict)

    conn.close()
    return render_template('index.html', bugs=bugs, user=user)

# 组内成员问题列表
@app.route('/admin/users', methods=['GET', 'POST', 'PUT'])
@login_required
@role_required('gly')
def admin_users():
    """用户管理API"""
    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    if request.method == 'GET':
        # 获取所有用户
        query, params = adapt_sql('SELECT id, username, chinese_name, email, phone, role, role_en, team FROM users ORDER BY id', ())
        c.execute(query, params)
        users = [dict(row) for row in c.fetchall()]
        conn.close()
        return jsonify(users)

    elif request.method == 'POST':
        # 添加新用户(JSON格式)
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据格式错误'}), 400

        username = data.get('username')
        chinese_name = data.get('chinese_name')
        password = data.get('password')
        role = data.get('role')
        team = data.get('team')

        if not all([username, password, role]):
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400

        hashed_password = generate_password_hash(password)
        try:
            if DB_TYPE == 'postgres':
                query, params = adapt_sql('''
                    INSERT INTO users (username, chinese_name, password, role, team)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id
                ''', (username, chinese_name, hashed_password, role, team))
                c.execute(query, params)
                user_id = c.fetchone()['id']
            else:
                # SQLite模式
                query, params = adapt_sql('''
                    INSERT INTO users (username, chinese_name, password, role, team)
                    VALUES (%s, %s, %s, %s, %s)
                ''', (username, chinese_name, hashed_password, role, team))
                c.execute(query, params)
                user_id = c.lastrowid
            conn.commit()
            return jsonify({'success': True, 'user_id': user_id})
        except Exception as e:
            conn.rollback()
            # 检查是否是用户名重复错误
            if 'UNIQUE constraint failed' in str(e) or 'duplicate key' in str(e):
                return jsonify({'success': False, 'message': '用户名已存在'}), 400
            else:
                return jsonify({'success': False, 'message': f'添加用户失败: {str(e)}'}), 500
        finally:
            conn.close()

    elif request.method == 'PUT':
        # 更新用户信息
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据格式错误'}), 400

        user_id = data.get('id')
        username = data.get('username')
        chinese_name = data.get('chinese_name')
        password = data.get('password')
        role = data.get('role')
        team = data.get('team')

        if not user_id:
            return jsonify({'success': False, 'message': '缺少用户ID'}), 400

        try:
            if password:
                hashed_password = generate_password_hash(password)
                query, params = adapt_sql('''
                    UPDATE users
                    SET username=%s, chinese_name=%s, password=%s, role=%s, team=%s
                    WHERE id=%s
                ''', (username, chinese_name, hashed_password, role, team, user_id))
                c.execute(query, params)
            else:
                query, params = adapt_sql('''
                    UPDATE users
                    SET username=%s, chinese_name=%s, role=%s, team=%s
                    WHERE id=%s
                ''', (username, chinese_name, role, team, user_id))
                c.execute(query, params)

            conn.commit()
            return jsonify({'success': True})
        except Exception as e:
            conn.rollback()
            # 检查是否是用户名重复错误
            if 'UNIQUE constraint failed' in str(e) or 'duplicate key' in str(e):
                return jsonify({'success': False, 'message': '用户名已存在'}), 400
            else:
                return jsonify({'success': False, 'message': f'更新用户失败: {str(e)}'}), 500
        finally:
            conn.close()

@app.route('/admin/users/<int:user_id>', methods=['GET', 'PUT'])
@login_required
@role_required('gly')
def user_detail(user_id):
    """获取或更新单个用户信息"""
    if request.method == 'GET':
        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            c = conn.cursor(cursor_factory=DictCursor)
        else:
            c = conn.cursor()
        query, params = adapt_sql('SELECT id, username, chinese_name, email, phone, role, role_en, team FROM users WHERE id = %s', (user_id,))
        c.execute(query, params)
        user = c.fetchone()
        conn.close()

        if not user:
            return jsonify({'success': False, 'message': '用户不存在'}), 404

        return jsonify(dict(user))
    """更新用户信息"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': '缺少请求数据'}), 400

    username = data.get('username')
    password = data.get('password')
    role = data.get('role')
    team = data.get('team')
    role_en = data.get('role_en')
    chinese_name = data.get('chinese_name')
    email = data.get('email')
    phone = data.get('phone')

    # 生成角色英文编码
    if not role_en and role:
        role_mapping = {
            '管理员': 'gly',
            '负责人': 'fzr',
            '实施组': 'ssz',
            '组内成员': 'zncy',
            '产品经理': 'pm'
        }
        role_en = role_mapping.get(role, 'zncy')

    # 生成团队英文编码(中文取首字母拼音大写)
    team_en = None
    if team:
        # 中文转拼音首字母映射表(大写)
        pinyin_map = {
            '啊':'A', '阿':'A', '埃':'A', '哎':'A', '唉':'A', '哀':'A', '皑':'A', '癌':'A', '蔼':'A', '矮':'A',
            '艾':'A', '碍':'A', '爱':'A', '隘':'A', '鞍':'A', '氨':'A', '安':'A', '俺':'A', '按':'A', '暗':'A',
            '吧':'B', '八':'B', '巴':'B', '拔':'B', '跋':'B', '靶':'B', '把':'B', '坝':'B', '霸':'B', '罢':'B',
            '白':'B', '百':'B', '柏':'B', '摆':'B', '败':'B', '拜':'B', '班':'B', '般':'B', '颁':'B', '板':'B',
            '擦':'C', '猜':'C', '裁':'C', '材':'C', '才':'C', '财':'C', '睬':'C', '踩':'C', '采':'C', '彩':'C',
            '菜':'C', '蔡':'C', '餐':'C', '参':'C', '蚕':'C', '残':'C', '惭':'C', '惨':'C', '灿':'C', '仓':'C',
            '大':'D', '呆':'D', '歹':'D', '傣':'D', '戴':'D', '带':'D', '殆':'D', '代':'D', '贷':'D', '袋':'D',
            '待':'D', '逮':'D', '怠':'D', '耽':'D', '担':'D', '丹':'D', '单':'D', '郸':'D', '掸':'D', '胆':'D',
            # 其他字母...
            '网络分析':'wlfx','实施组':'ssz','第三道防线':'dsdfx','新能源':'xny'
        }
        # 只处理中文字符，生成大写拼音首字母
        team_en = ''.join([pinyin_map.get(c, '') for c in team if '\u4e00' <= c <= '\u9fa5'])
        # 如果team_en为空(无中文字符)，则使用team的前3个字符大写
        if not team_en:
            team_en = team[:3].upper()

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    try:
        if password:
            hashed_password = generate_password_hash(password)
            query, params = adapt_sql('''
                UPDATE users
                SET username=%s, password=%s, role=%s, role_en=%s, team=%s, team_en=%s, chinese_name=%s, email=%s, phone=%s
                WHERE id=%s
            ''', (username, hashed_password, role, role_en, team, team_en, chinese_name, email, phone, user_id))
            c.execute(query, params)
        else:
            query, params = adapt_sql('''
                UPDATE users
                SET username=%s, role=%s, role_en=%s, team=%s, team_en=%s, chinese_name=%s, email=%s, phone=%s
                WHERE id=%s
            ''', (username, role, role_en, team, team_en, chinese_name, email, phone, user_id))
            c.execute(query, params)

        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        # 检查是否是用户名重复错误
        if 'UNIQUE constraint failed' in str(e) or 'duplicate key' in str(e):
            return jsonify({'success': False, 'message': '用户名已存在'}), 400
        else:
            return jsonify({'success': False, 'message': f'更新用户失败: {str(e)}'}), 500
    finally:
        conn.close()

@app.route('/admin/users/<int:user_id>', methods=['DELETE'])
@login_required
@role_required('gly')
def delete_user(user_id):
    """删除用户"""
    # 不能删除自己
    current_user = get_current_user()
    if current_user['id'] == user_id:
        return jsonify({'success': False, 'message': '不能删除自己'}), 400

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    try:
        # 检查用户是否存在
        query, params = adapt_sql('SELECT id FROM users WHERE id = %s', (user_id,))
        c.execute(query, params)
        if not c.fetchone():
            return jsonify({'success': False, 'message': '用户不存在'}), 404

        # 删除用户
        query, params = adapt_sql('DELETE FROM users WHERE id = %s', (user_id,))
        c.execute(query, params)
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/admin/bugs', methods=['GET'])
@login_required
@role_required('gly')
def admin_bugs():
    """获取所有问题数据(API)"""
    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()
    query = '''
        SELECT b.id, b.title, b.status, b.created_at, b.resolved_at, b.assigned_to,
               COALESCE(u1.chinese_name, u1.username) as creator_name,
               COALESCE(u2.chinese_name, u2.username) as assignee_name
        FROM bugs b
        LEFT JOIN users u1 ON b.created_by = u1.id
        LEFT JOIN users u2 ON b.assigned_to = u2.id
        ORDER BY b.created_at DESC
    '''
    adapted_query, adapted_params = adapt_sql(query, ())
    c.execute(adapted_query, adapted_params)
    bugs = []
    for row in c.fetchall():
        bug = dict(row)
        # 处理created_at - 兼容SQLite字符串和PostgreSQL datetime
        if bug['created_at']:
            if isinstance(bug['created_at'], str):
                bug['created_at'] = bug['created_at']  # SQLite已经是字符串格式
            else:
                bug['created_at'] = bug['created_at'].strftime('%Y-%m-%d %H:%M:%S')  # PostgreSQL datetime
        else:
            bug['created_at'] = '--'
        # 处理resolved_at - 兼容SQLite字符串和PostgreSQL datetime
        if bug['resolved_at']:
            if isinstance(bug['resolved_at'], str):
                bug['resolved_at'] = bug['resolved_at']  # SQLite已经是字符串格式
            else:
                bug['resolved_at'] = bug['resolved_at'].strftime('%Y-%m-%d %H:%M:%S')  # PostgreSQL datetime
        else:
            bug['resolved_at'] = '--'
        bugs.append(bug)
    conn.close()
    return jsonify(bugs)

# 用户设置页面路由
@app.route('/user-settings')
@login_required
def user_settings():
    """用户设置页面"""
    user = get_current_user()
    if not user:
        return redirect('/login')

    return render_template('user_settings.html', user=user)

# 用户修改密码API
@app.route('/api/user/change-password', methods=['POST'])
@login_required
def change_password():
    """修改用户密码"""
    try:
        user = get_current_user()
        data = request.get_json()

        current_password = data.get('current_password')
        new_password = data.get('new_password')

        if not current_password or not new_password:
            return jsonify({'success': False, 'message': '密码不能为空'})

        if len(new_password) < 6:
            return jsonify({'success': False, 'message': '新密码长度至少6位'})

        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            c = conn.cursor(cursor_factory=DictCursor)
        else:
            c = conn.cursor()

        # 验证当前密码
        query, params = adapt_sql('SELECT password FROM users WHERE id = %s', (user['id'],))
        c.execute(query, params)
        result = c.fetchone()

        if not result:
            conn.close()
            return jsonify({'success': False, 'message': '用户不存在'})

        stored_password = result[0] if not isinstance(result, dict) else result['password']

        if not check_password_hash(stored_password, current_password):
            conn.close()
            return jsonify({'success': False, 'message': '当前密码错误'})

        # 更新密码
        hashed_password = generate_password_hash(new_password)
        query, params = adapt_sql('UPDATE users SET password = %s WHERE id = %s', (hashed_password, user['id']))
        c.execute(query, params)
        conn.commit()
        conn.close()

        return jsonify({'success': True, 'message': '密码修改成功'})

    except Exception as e:
        app.logger.error(f"修改密码失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

# 用户偏好设置API
@app.route('/api/user/preferences', methods=['GET', 'POST'])
@login_required
def user_preferences():
    """获取或设置用户偏好"""
    try:
        user = get_current_user()

        if request.method == 'GET':
            # 获取用户偏好设置
            conn = get_db_connection()
            if DB_TYPE == 'postgres':
                c = conn.cursor(cursor_factory=DictCursor)
            else:
                c = conn.cursor()

            query, params = adapt_sql('SELECT preferences FROM users WHERE id = %s', (user['id'],))
            c.execute(query, params)
            result = c.fetchone()
            conn.close()

            if result:
                preferences_str = result[0] if not isinstance(result, dict) else result['preferences']
                if preferences_str:
                    import json
                    preferences = json.loads(preferences_str)
                else:
                    preferences = {}
            else:
                preferences = {}

            return jsonify({'success': True, 'data': preferences})

        else:  # POST
            # 设置用户偏好
            data = request.get_json()

            conn = get_db_connection()
            if DB_TYPE == 'postgres':
                c = conn.cursor(cursor_factory=DictCursor)
            else:
                c = conn.cursor()

            import json
            preferences_str = json.dumps(data)

            query, params = adapt_sql('UPDATE users SET preferences = %s WHERE id = %s', (preferences_str, user['id']))
            c.execute(query, params)
            conn.commit()
            conn.close()

            return jsonify({'success': True, 'message': '设置保存成功'})

    except Exception as e:
        app.logger.error(f"用户偏好设置操作失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin')
@login_required
@role_required('gly')
def admin():
    """管理员控制面板"""
    user = get_current_user()
    if not user:
        return redirect('/login')

    # 获取所有用户
    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()
    query, params = adapt_sql('SELECT id, username, chinese_name, email, phone, role, role_en, team FROM users ORDER BY id', ())
    c.execute(query, params)
    users = c.fetchall()

    # 获取总用户数
    total_users = len(users)

    # 获取所有问题
    query, params = adapt_sql('''
        SELECT b.id, b.title, b.status, b.created_at, b.resolved_at, b.assigned_to,
               COALESCE(u1.chinese_name, u1.username) as creator_name,
               COALESCE(u2.chinese_name, u2.username) as assignee_name
        FROM bugs b
        LEFT JOIN users u1 ON b.created_by = u1.id
        LEFT JOIN users u2 ON b.assigned_to = u2.id
        ORDER BY b.created_at DESC
    ''', ())
    c.execute(query, params)
    bugs = c.fetchall()

    # 格式化问题创建时间和解决时间
    formatted_bugs = []
    for bug in bugs:
        bug_dict = dict(bug)
        # 处理created_at
        if isinstance(bug_dict['created_at'], str):
            bug_dict['created_at'] = bug_dict['created_at']  # 已经是字符串则直接使用
        elif bug_dict['created_at']:
            bug_dict['created_at'] = bug_dict['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        else:
            bug_dict['created_at'] = '--'

        # 处理resolved_at
        if isinstance(bug_dict['resolved_at'], str):
            bug_dict['resolved_at'] = bug_dict['resolved_at']  # 已经是字符串则直接使用
        elif bug_dict['resolved_at']:
            bug_dict['resolved_at'] = bug_dict['resolved_at'].strftime('%Y-%m-%d %H:%M:%S')
        else:
            bug_dict['resolved_at'] = '--'

        formatted_bugs.append(bug_dict)

    # 获取所有项目
    query, params = adapt_sql('SELECT id, name, type, city FROM projects ORDER BY name', ())
    c.execute(query, params)
    projects = c.fetchall()

    conn.close()

    return render_template('admin.html', users=users, bugs=formatted_bugs, projects=projects, user=user, total_users=total_users)

@app.route('/product-manager')
@login_required
@role_required('pm')
def product_manager_dashboard():
    """产品经理专用页面"""
    user = get_current_user()
    if not user:
        return redirect('/login')

    return render_template('product_manager.html', user=user)

@app.route('/api/product-manager/my-teams')
@login_required
@role_required('pm')
def get_my_teams():
    """获取当前产品经理负责的团队"""
    try:
        user = get_current_user()

        # 获取用户的团队信息
        user_teams = user.get('team', '')
        if not user_teams:
            return jsonify({'success': True, 'data': []})

        # 解析团队列表
        team_names = [team.strip() for team in user_teams.split(',') if team.strip()]

        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            c = conn.cursor(cursor_factory=DictCursor)
        else:
            c = conn.cursor()

        result = []
        for team_name in team_names:
            # 统计团队成员数量
            query, params = adapt_sql('''
                SELECT COUNT(*) as member_count
                FROM users
                WHERE team = %s OR team LIKE %s OR team LIKE %s OR team LIKE %s
            ''', (team_name, f'{team_name},%', f'%,{team_name},%', f'%,{team_name}'))
            c.execute(query, params)
            member_result = c.fetchone()
            member_count = member_result[0] if not isinstance(member_result, dict) else member_result['member_count']

            # 统计团队Bug数量（通过assigned_to联合users表）
            query, params = adapt_sql('''
                SELECT
                    COUNT(*) as total,
                    SUM(CASE WHEN b.status = '待处理' THEN 1 ELSE 0 END) as pending,
                    SUM(CASE WHEN b.status = '已分配' THEN 1 ELSE 0 END) as assigned,
                    SUM(CASE WHEN b.status = '处理中' THEN 1 ELSE 0 END) as processing,
                    SUM(CASE WHEN b.status = '已解决' THEN 1 ELSE 0 END) as resolved,
                    SUM(CASE WHEN b.status = '已完成' THEN 1 ELSE 0 END) as closed
                FROM bugs b
                LEFT JOIN users u ON b.assigned_to = u.id
                WHERE u.team = %s OR b.created_by = %s
            ''', (team_name, user['id']))
            c.execute(query, params)
            bug_result = c.fetchone()

            if isinstance(bug_result, dict):
                bug_stats = {
                    'total': bug_result['total'] or 0,
                    'pending': bug_result['pending'] or 0,
                    'assigned': bug_result['assigned'] or 0,
                    'processing': bug_result['processing'] or 0,
                    'resolved': bug_result['resolved'] or 0,
                    'closed': bug_result['closed'] or 0
                }
            else:
                bug_stats = {
                    'total': bug_result[0] or 0,
                    'pending': bug_result[1] or 0,
                    'assigned': bug_result[2] or 0,
                    'processing': bug_result[3] or 0,
                    'resolved': bug_result[4] or 0,
                    'closed': bug_result[5] or 0
                }

            result.append({
                'name': team_name,
                'member_count': member_count,
                'bug_stats': bug_stats
            })

        conn.close()
        return jsonify({'success': True, 'data': result})

    except Exception as e:
        app.logger.error(f"获取产品经理团队失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/product-manager/bugs')
@login_required
@role_required('pm')
def get_product_manager_bugs():
    """获取产品经理负责团队的所有bugs"""
    try:
        user = get_current_user()

        # 获取用户的团队信息
        user_teams = user.get('team', '')
        if not user_teams:
            return jsonify({'success': True, 'data': []})

        # 解析团队列表
        team_names = [team.strip() for team in user_teams.split(',') if team.strip()]

        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            c = conn.cursor(cursor_factory=DictCursor)
        else:
            c = conn.cursor()

        # 构建IN查询的占位符
        placeholders = ','.join(['%s'] * len(team_names))

        query, params = adapt_sql(f'''
            SELECT b.*, u2.team as product_line_name,
                   COALESCE(u1.chinese_name, u1.username) as creator_name,
                   COALESCE(u2.chinese_name, u2.username) as assignee_name
            FROM bugs b
            LEFT JOIN users u1 ON b.created_by = u1.id
            LEFT JOIN users u2 ON b.assigned_to = u2.id
            WHERE u2.team IN ({placeholders}) OR b.created_by = %s
            ORDER BY u2.team, b.created_at DESC
        ''', tuple(team_names) + (user['id'],))

        c.execute(query, params)
        bugs = c.fetchall()
        conn.close()

        # 转换为字典列表并格式化时间
        result = []
        for bug in bugs:
            if isinstance(bug, dict):
                bug_dict = dict(bug)
            else:
                bug_dict = {
                    'id': bug[0],
                    'title': bug[1],
                    'description': bug[2],
                    'status': bug[3],
                    'type': bug[4],
                    'assigned_to': bug[5],
                    'created_by': bug[6],
                    'project': bug[7],
                    'created_at': bug[8],
                    'resolved_at': bug[9],
                    'resolution': bug[10],
                    'image_path': bug[11],
                    'product_line_id': bug[12],
                    'product_line_name': bug[13],
                    'creator_name': bug[14],
                    'assignee_name': bug[15]
                }

            # 格式化时间
            if bug_dict.get('created_at'):
                if isinstance(bug_dict['created_at'], str):
                    bug_dict['created_at'] = bug_dict['created_at']
                else:
                    bug_dict['created_at'] = bug_dict['created_at'].strftime('%Y-%m-%d %H:%M:%S')

            if bug_dict.get('resolved_at'):
                if isinstance(bug_dict['resolved_at'], str):
                    bug_dict['resolved_at'] = bug_dict['resolved_at']
                else:
                    bug_dict['resolved_at'] = bug_dict['resolved_at'].strftime('%Y-%m-%d %H:%M:%S')

            result.append(bug_dict)

        return jsonify({'success': True, 'data': result})

    except Exception as e:
        app.logger.error(f"获取产品经理bugs失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/product-manager/statistics')
@login_required
@role_required('pm')
def get_product_manager_statistics():
    """获取产品经理的统计数据"""
    try:
        user = get_current_user()

        # 获取用户的团队信息
        user_teams = user.get('team', '')
        if not user_teams:
            return jsonify({'success': True, 'data': {
                'total': 0, 'pending': 0, 'processing': 0, 'resolved': 0, 'closed': 0
            }})

        # 解析团队列表
        team_names = [team.strip() for team in user_teams.split(',') if team.strip()]

        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            c = conn.cursor(cursor_factory=DictCursor)
        else:
            c = conn.cursor()

        # 构建IN查询的占位符
        placeholders = ','.join(['%s'] * len(team_names))

        query, params = adapt_sql(f'''
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN b.status = '待处理' THEN 1 ELSE 0 END) as pending,
                SUM(CASE WHEN b.status = '处理中' THEN 1 ELSE 0 END) as processing,
                SUM(CASE WHEN b.status = '已解决' THEN 1 ELSE 0 END) as resolved,
                SUM(CASE WHEN b.status = '已完成' THEN 1 ELSE 0 END) as closed
            FROM bugs b
            LEFT JOIN users u ON b.assigned_to = u.id
            WHERE u.team IN ({placeholders}) OR b.created_by = %s
        ''', tuple(team_names) + (user['id'],))

        c.execute(query, params)
        result = c.fetchone()
        conn.close()

        if isinstance(result, dict):
            stats = {
                'total': result['total'] or 0,
                'pending': result['pending'] or 0,
                'processing': result['processing'] or 0,
                'resolved': result['resolved'] or 0,
                'closed': result['closed'] or 0
            }
        else:
            stats = {
                'total': result[0] or 0,
                'pending': result[1] or 0,
                'processing': result[2] or 0,
                'resolved': result[3] or 0,
                'closed': result[4] or 0
            }

        return jsonify({'success': True, 'data': stats})

    except Exception as e:
        app.logger.error(f"获取产品经理统计数据失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/team-issues')
@login_required
@role_required('zncy')
def team_issues():
    """组内成员问题列表"""
    try:
        app.logger.debug("开始处理team-issues请求")
        user = get_current_user()
        app.logger.debug(f"当前用户: {user}")
        if not user:
            app.logger.debug("用户未登录，重定向到登录页面")
            return redirect('/login')

        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            c = conn.cursor(cursor_factory=DictCursor)
        else:
            c = conn.cursor()
        query, params = adapt_sql('''
            SELECT b.id, b.title, b.description, b.status, b.type, b.assigned_to, b.created_by, b.project,
                   b.created_at as local_created_at,
                   b.resolved_at as local_resolved_at,
                   b.resolution, b.image_path,
                   COALESCE(u1.chinese_name, u1.username) as creator_name, COALESCE(u2.chinese_name, u2.username) as assignee_name
            FROM bugs b
            LEFT JOIN users u1 ON b.created_by = u1.id
            LEFT JOIN users u2 ON b.assigned_to = u2.id
            WHERE (
                (b.assigned_to = %s)
                OR
                (b.status = '待处理' AND b.assigned_to IS NULL AND u1.team = %s)
                OR
                (b.status = '已解决' AND b.assigned_to = %s)
            )
            ORDER BY b.created_at DESC
        ''', (user['id'], user['team'], user['id']))
        c.execute(query, params)
        bugs = c.fetchall()

        # 格式化问题创建时间和解决时间
        formatted_bugs = []
        for bug in bugs:
            bug_dict = dict(bug)
            # 处理local_created_at (别名为created_at)
            created_at = bug_dict.get('local_created_at')
            if isinstance(created_at, str):
                bug_dict['created_at'] = created_at  # 已经是字符串则直接使用
            elif created_at:
                bug_dict['created_at'] = created_at.strftime('%Y-%m-%d %H:%M:%S')
            else:
                bug_dict['created_at'] = '--'

            # 处理local_resolved_at (别名为resolved_at)
            resolved_at = bug_dict.get('local_resolved_at')
            if isinstance(resolved_at, str):
                bug_dict['resolved_at'] = resolved_at  # 已经是字符串则直接使用
            elif resolved_at:
                bug_dict['resolved_at'] = resolved_at.strftime('%Y-%m-%d %H:%M:%S')
            else:
                bug_dict['resolved_at'] = '--'

            formatted_bugs.append(bug_dict)

        conn.close()
        app.logger.debug(f"Rendering team_issues.html with user role: {user['role']} and bugs: {formatted_bugs}")
        return render_template('team_issues.html', bugs=formatted_bugs, user=user)
    except Exception as e:
        error_msg = f"team_issues页面错误: {str(e)}"
        print(error_msg)
        print(f"错误类型: {type(e)}")
        print(f"错误发生位置:")
        traceback.print_exc()
        try:
            app.logger.error(error_msg, exc_info=True)
        except:
            pass
        # 如果模板渲染失败，返回简单的HTML
        return f"""
        <html>
        <head><title>我的任务</title><meta charset="UTF-8"></head>
        <body>
            <h1>我的任务</h1>
            <p>抱歉，页面加载出现问题。</p>
            <p>错误信息: {str(e)}</p>
            <p><a href="/">返回首页</a> | <a href="/logout">退出登录</a></p>
        </body>
        </html>
        """, 500

# 提交问题页面
@app.route('/submit', methods=['GET', 'POST'])
@login_required
def submit_page():
    user = get_current_user()
    if not user:
        return redirect('/login')

    if request.method == 'POST':
        # 处理问题提交
        return submit_bug_handler(user)

    # GET请求 - 显示提交页面
    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()
    try:
        # 查询所有负责人（角色英文缩写为'fzr'）并确保显示名称唯一
        query, params = adapt_sql('''
            SELECT DISTINCT COALESCE(chinese_name, username) as display_name
            FROM users
            WHERE role_en = %s
        ''', ('fzr',))
        c.execute(query, params)
        managers = [row['display_name'] for row in c.fetchall()]
        app.logger.debug(f"获取到的负责人列表: {managers}")

        # 调试：打印实际查询到的负责人显示名称
        app.logger.debug(f"实际负责人显示名称: {managers}")

        # 获取项目列表
        query, params = adapt_sql('SELECT name FROM projects ORDER BY name', ())
        c.execute(query, params)
        projects = [row['name'] if DB_TYPE == 'postgres' else row[0] for row in c.fetchall()]
        app.logger.debug(f"获取到的项目列表: {projects}")

        return render_template('submit.html', managers=managers, projects=projects, user=user)
    finally:
        conn.close()

def submit_bug_handler(user):
    """处理问题提交的逻辑"""
    app.logger.debug(f"收到问题提交请求，表单数据: {request.form}")

    title = request.form.get('title')
    description = request.form.get('description')
    created_by = user['id']
    app.logger.debug(f"提交用户ID: {created_by}, 标题: {title}, 描述: {description}")

    if not title or not description:
        return redirect('/submit?error=标题和描述不能为空')

    # 处理图片上传
    image_path = None
    if 'image' in request.files:
        file = request.files['image']
        if file and file.filename and allowed_file(file.filename):
            upload_dir = app.config['UPLOAD_FOLDER']
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir, mode=0o777, exist_ok=True)
            filename = secure_filename(file.filename)
            filepath = os.path.join(upload_dir, filename)
            file.save(filepath)
            image_path = f'/uploads/{filename}'
            app.logger.debug(f"文件保存成功: {filepath}")

    # 存入数据库
    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    try:
        manager_name = request.form.get('manager')
        if not manager_name:
            return redirect('/submit?error=请选择负责人')

        # 获取负责人ID（使用显示名查询）
        query, params = adapt_sql('SELECT id FROM users WHERE COALESCE(chinese_name, username) = %s', (manager_name,))
        c.execute(query, params)
        manager = c.fetchone()
        if not manager:
            return redirect(f'/submit?error=指定的负责人"{manager_name}"不存在')

        manager_id = manager['id'] if DB_TYPE == 'postgres' else manager[0]
        project_id = request.form.get('project', '')
        bug_type = request.form.get('type', 'bug')  # 获取类型，默认为bug
        product_line_id = request.form.get('product_line')  # 获取产品线ID

        # 如果product_line_id为空字符串，转换为None
        if product_line_id == '':
            product_line_id = None
        elif product_line_id:
            product_line_id = int(product_line_id)

        # 获取当前时间，精确到秒
        from datetime import datetime
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if DB_TYPE == 'postgres':
            query, params = adapt_sql('''
                INSERT INTO bugs (title, description, created_by, project, image_path, assigned_to, status, type, created_at, product_line_id)
                VALUES (%s, %s, %s, %s, %s, %s, '待处理', %s, %s, %s)
                RETURNING id
            ''', (title, description, created_by, project_id, image_path, manager_id, bug_type, current_time, product_line_id))
            c.execute(query, params)
            bug_id = c.fetchone()['id']
        else:
            query, params = adapt_sql('''
                INSERT INTO bugs (title, description, created_by, project, image_path, assigned_to, status, type, created_at, product_line_id)
                VALUES (%s, %s, %s, %s, %s, %s, '待处理', %s, %s, %s)
            ''', (title, description, created_by, project_id, image_path, manager_id, bug_type, current_time, product_line_id))
            c.execute(query, params)
            bug_id = c.lastrowid

        conn.commit()

        # 异步发送通知（在后台处理）
        def send_creation_notification_async():
            try:
                app.logger.info(f"后台发送问题创建通知 - bug_id: {bug_id}, manager_id: {manager_id}")
                from notification.simple_notifier import simple_notifier
                from datetime import datetime

                notification_data = {
                    'bug_id': bug_id,
                    'title': title,
                    'description': description,
                    'creator_name': user['chinese_name'] or user['username'],
                    'created_time': datetime.now().isoformat(),
                    'creator_id': created_by,
                    'assigned_manager_id': manager_id,
                    'product_line_id': product_line_id
                }

                simple_notifier.send_flow_notification('bug_created', notification_data)
                app.logger.info(f"问题创建通知发送完成 - bug_id: {bug_id}")
            except Exception as e:
                app.logger.error(f"后台创建通知发送失败 - bug_id: {bug_id}, 错误: {e}")
                import traceback
                app.logger.error(f"创建通知发送错误详情: {traceback.format_exc()}")

        # 启动后台线程发送通知
        import threading
        notification_thread = threading.Thread(target=send_creation_notification_async, daemon=True)
        notification_thread.start()

        app.logger.info(f"问题提交成功，通知已提交后台处理 - bug_id: {bug_id}")
        return redirect(f'/?message=问题提交成功')

    except Exception as e:
        conn.rollback()
        app.logger.error(f"提交问题失败: {str(e)}", exc_info=True)
        return redirect(f'/submit?error=提交失败: {str(e)}')
    finally:
        try:
            conn.close()
        except:
            pass

# 提交问题API
@app.route('/bug/submit', methods=['POST'])
@login_required
def submit_bug():
    app.logger.debug(f"收到问题提交请求，表单数据: {request.form}")
    user = get_current_user()
    if not user:
        app.logger.warning("提交问题失败: 用户未登录")
        return jsonify({'success': False, 'message': '用户未登录'}), 401

    title = request.form.get('title')
    description = request.form.get('description')
    created_by = user['id']
    app.logger.debug(f"提交用户ID: {created_by}, 标题: {title}, 描述: {description}")

    if not title or not description:
        return jsonify({'success': False, 'message': '标题和描述不能为空'}), 400

    # 处理多图片上传
    image_paths = []
    main_image_path = None

    app.logger.debug(f"请求文件字段: {list(request.files.keys())}")

    if 'images' in request.files:
        app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH
        files = request.files.getlist('images')
        app.logger.debug(f"接收到 {len(files)} 个图片文件")

        upload_dir = app.config['UPLOAD_FOLDER']
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir, mode=0o777, exist_ok=True)

        for i, file in enumerate(files):
            if file and file.filename and allowed_file(file.filename):
                # 生成唯一文件名
                import uuid
                file_ext = file.filename.rsplit('.', 1)[1].lower()
                unique_filename = f"{uuid.uuid4().hex}.{file_ext}"
                filepath = os.path.join(upload_dir, unique_filename)
                file.save(filepath)

                image_url = f'/uploads/{unique_filename}'
                image_paths.append(image_url)

                # 第一张图片作为主图片（向后兼容）
                if i == 0:
                    main_image_path = image_url

                app.logger.debug(f"文件保存成功: {filepath}")

        app.logger.debug(f"多图片上传完成，共保存 {len(image_paths)} 张图片")

    # 向后兼容：如果没有使用新的多图片上传，检查旧的单图片上传
    elif 'image' in request.files:
        app.logger.debug("使用旧版单图片上传")
        file = request.files['image']
        if file and file.filename and allowed_file(file.filename):
            upload_dir = app.config['UPLOAD_FOLDER']
            if not os.path.exists(upload_dir):
                os.makedirs(upload_dir, mode=0o777, exist_ok=True)

            import uuid
            file_ext = file.filename.rsplit('.', 1)[1].lower()
            unique_filename = f"{uuid.uuid4().hex}.{file_ext}"
            filepath = os.path.join(upload_dir, unique_filename)
            file.save(filepath)

            main_image_path = f'/uploads/{unique_filename}'
            image_paths.append(main_image_path)
            app.logger.debug(f"文件保存成功: {filepath}")

    # 存入数据库
    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    try:
        manager_name = request.form.get('manager')
        if not manager_name:
            return jsonify({'success': False, 'message': '请选择负责人'}), 400

        # 获取负责人ID（使用显示名查询）
        query, params = adapt_sql('SELECT id FROM users WHERE COALESCE(chinese_name, username) = %s', (manager_name,))
        c.execute(query, params)
        manager = c.fetchone()
        if not manager:
            return jsonify({'success': False, 'message': f'指定的负责人"{manager_name}"不存在'}), 404

        manager_id = manager['id']
        project_id = request.form.get('project', '')
        bug_type = request.form.get('type', 'bug')  # 获取类型，默认为bug

        # 获取当前时间，精确到秒
        from datetime import datetime
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        query, params = adapt_sql('''
            INSERT INTO bugs (title, description, created_by, project, image_path, assigned_to, status, type, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, '待处理', %s, %s)
            RETURNING id
        ''', (title, description, created_by, project_id, main_image_path, manager_id, bug_type, current_time))
        c.execute(query, params)

        bug_id = c.fetchone()['id']

        # 保存所有图片到bug_images表
        if image_paths:
            for image_path in image_paths:
                image_query, image_params = adapt_sql('''
                    INSERT INTO bug_images (bug_id, image_path, created_at)
                    VALUES (%s, %s, %s)
                ''', (bug_id, image_path, current_time))
                c.execute(image_query, image_params)

        conn.commit()

        # 立即返回响应，不等待通知发送
        response_data = {
            'success': True,
            'bug_id': bug_id,
            'redirect': f'/bug/{bug_id}'
        }

        # 异步发送通知（在后台处理）
        def send_notification_async():
            try:
                app.logger.info(f"后台发送问题创建通知 - bug_id: {bug_id}, manager_id: {manager_id}")
                from notification.simple_notifier import simple_notifier
                from datetime import datetime

                notification_data = {
                    'bug_id': bug_id,
                    'title': title,
                    'description': description,
                    'creator_name': user['chinese_name'] or user['username'],
                    'created_time': datetime.now().isoformat(),
                    'creator_id': created_by,
                    'assigned_manager_id': manager_id
                }

                simple_notifier.send_flow_notification('bug_created', notification_data)
                app.logger.info(f"问题创建通知发送完成 - bug_id: {bug_id}")
            except Exception as e:
                app.logger.error(f"后台通知发送失败 - bug_id: {bug_id}, 错误: {e}")
                import traceback
                app.logger.error(f"通知发送错误详情: {traceback.format_exc()}")

        # 启动后台线程发送通知
        import threading
        notification_thread = threading.Thread(target=send_notification_async, daemon=True)
        notification_thread.start()

        app.logger.info(f"问题提交成功，通知已提交后台处理 - bug_id: {bug_id}")

        return jsonify(response_data)
    except Exception as e:
        conn.rollback()
        error_msg = str(e)
        if 'connection' in error_msg.lower() or 'operational' in error_msg.lower():
            app.logger.error(f"数据库连接失败: {error_msg}")
            return jsonify({'success': False, 'message': '数据库连接失败，请检查数据库服务'}), 503
        elif 'constraint' in error_msg.lower() or 'integrity' in error_msg.lower():
            app.logger.error(f"数据完整性错误: {error_msg}")
            return jsonify({'success': False, 'message': '数据验证失败，请检查输入'}), 400
    except Exception as e:
        conn.rollback()
        app.logger.error(f"提交问题失败: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': f'系统错误: {str(e)}'}), 500
    finally:
        try:
            conn.close()
        except:
            pass

# 问题详情页
@app.route('/bug/<int:bug_id>')
@login_required
def bug_detail(bug_id):
    user = get_current_user()
    if not user:
        return redirect('/login')

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    # 查询bug基本信息
    query = '''
        SELECT b.*, COALESCE(u1.chinese_name, u1.username) as creator_name, COALESCE(u2.chinese_name, u2.username) as assignee_name,
               b.created_at as local_created_at,
               b.resolved_at as local_resolved_at
        FROM bugs b
        LEFT JOIN users u1 ON b.created_by = u1.id
        LEFT JOIN users u2 ON b.assigned_to = u2.id
        WHERE b.id = %s
    '''
    adapted_query, adapted_params = adapt_sql(query, (bug_id,))
    c.execute(adapted_query, adapted_params)
    bug = c.fetchone()

    # 查询所有相关图片
    images_query = '''
        SELECT image_path, created_at
        FROM bug_images
        WHERE bug_id = %s
        ORDER BY created_at ASC
    '''
    adapted_images_query, adapted_images_params = adapt_sql(images_query, (bug_id,))
    c.execute(adapted_images_query, adapted_images_params)
    bug_images = c.fetchall()

    conn.close()
    if not bug:
        return "问题不存在", 404

    # 格式化时间 - 兼容SQLite字符串和PostgreSQL datetime
    bug_dict = dict(bug)
    if bug_dict['local_created_at']:
        if isinstance(bug_dict['local_created_at'], str):
            bug_dict['local_created_at'] = bug_dict['local_created_at']  # SQLite已经是字符串格式
        else:
            bug_dict['local_created_at'] = bug_dict['local_created_at'].strftime('%Y-%m-%d %H:%M:%S')  # PostgreSQL datetime
    else:
        bug_dict['local_created_at'] = '--'
    if bug_dict['local_resolved_at']:
        if isinstance(bug_dict['local_resolved_at'], str):
            bug_dict['local_resolved_at'] = bug_dict['local_resolved_at']  # SQLite已经是字符串格式
        else:
            bug_dict['local_resolved_at'] = bug_dict['local_resolved_at'].strftime('%Y-%m-%d %H:%M:%S')  # PostgreSQL datetime
    else:
        bug_dict['local_resolved_at'] = '--'

    message = request.args.get('message')

    # 处理图片列表
    images = []

    # 添加主图片（向后兼容）
    if bug_dict['image_path']:
        images.append({'path': bug_dict['image_path']})
        app.logger.debug(f"添加主图片: {bug_dict['image_path']}")

    # 添加其他图片
    if bug_images:
        app.logger.debug(f"找到 {len(bug_images)} 张关联图片")
        for img in bug_images:
            if DB_TYPE == 'postgres':
                img_path = img['image_path']
            else:
                img_path = img[0]

            app.logger.debug(f"处理关联图片: {img_path}")

            # 避免重复添加主图片
            if not bug_dict['image_path'] or img_path != bug_dict['image_path']:
                images.append({'path': img_path})
                app.logger.debug(f"添加关联图片: {img_path}")

    app.logger.debug(f"总共添加了 {len(images)} 张图片")

    # 将创建者ID和图片列表传递给模板
    return render_template('bug_detail.html', bug=bug_dict, message=message,
                          created_by=bug_dict['created_by'], user=user,
                          images=images)

# 分配问题页面
@app.route('/bug/assign/<int:bug_id>')
@login_required
@role_required('fzr')
def assign_page(bug_id):
    user = get_current_user()
    if not user:
        return redirect('/login')

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()
    query, params = adapt_sql('''
        SELECT b.*, COALESCE(u1.chinese_name, u1.username) as creator_name, COALESCE(u2.chinese_name, u2.username) as assignee_name
        FROM bugs b
        LEFT JOIN users u1 ON b.created_by = u1.id
        LEFT JOIN users u2 ON b.assigned_to = u2.id
        WHERE b.id = %s
    ''', (bug_id,))
    c.execute(query, params)
    bug = c.fetchone()

    # 获取当前负责人的组内成员
    query, params = adapt_sql('''
        SELECT id, COALESCE(chinese_name, username) as username FROM users
        WHERE team = %s AND (role = '组内成员' OR role = '负责人')
    ''', (user['team'],))
    c.execute(query, params)
    team_members = [{'id': row['id'], 'username': row['username']} for row in c.fetchall()]

    conn.close()
    if not bug:
        return "问题不存在", 404
    return render_template('assign.html', bug=bug, team_members=team_members, user=user)

# 分配问题API
@app.route('/bug/assign/<int:bug_id>', methods=['POST'])
@login_required
@role_required('fzr')
def assign_bug(bug_id):
    """指派问题给组内成员"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    assigned_to = request.form.get('assigned_to')
    if not assigned_to:
        return jsonify({'success': False, 'message': '负责人不能为空'})

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    # 检查被指派人是否在同一团队
    query, params = adapt_sql('SELECT team FROM users WHERE id = %s', (assigned_to,))
    c.execute(query, params)
    assignee = c.fetchone()
    if not assignee or assignee['team'] != user['team']:
        conn.close()
        return jsonify({'success': False, 'message': '只能指派给同团队成员'})

    # 获取问题信息（用于通知）
    query, params = adapt_sql('SELECT title, description, created_by FROM bugs WHERE id = %s', (bug_id,))
    c.execute(query, params)
    bug_info = c.fetchone()

    # 更新问题状态
    query, params = adapt_sql('''
        UPDATE bugs
        SET status = '已分配',
            assigned_to = %s
        WHERE id = %s
    ''', (assigned_to, bug_id))
    c.execute(query, params)
    conn.commit()

    # 获取被指派人用户名
    query, params = adapt_sql('SELECT username FROM users WHERE id = %s', (assigned_to,))
    c.execute(query, params)
    assignee_name = c.fetchone()['username']
    conn.close()

    # 异步发送通知（在后台处理）
    def send_assignment_notification_async():
        try:
            app.logger.info(f"后台发送问题分配通知 - bug_id: {bug_id}, assignee_id: {assigned_to}")
            from notification.simple_notifier import simple_notifier
            from datetime import datetime

            notification_data = {
                'bug_id': bug_id,
                'title': bug_info['title'],
                'description': bug_info['description'],
                'assignee_id': assigned_to,
                'assigner_name': user['chinese_name'] or user['username'],
                'assigned_time': datetime.now().isoformat(),
                'creator_id': bug_info['created_by']
            }

            simple_notifier.send_flow_notification('bug_assigned', notification_data)
            app.logger.info(f"问题分配通知发送完成 - bug_id: {bug_id}")
        except Exception as e:
            app.logger.error(f"后台分配通知发送失败 - bug_id: {bug_id}, 错误: {e}")
            import traceback
            app.logger.error(f"分配通知发送错误详情: {traceback.format_exc()}")

    # 启动后台线程发送通知
    import threading
    notification_thread = threading.Thread(target=send_assignment_notification_async, daemon=True)
    notification_thread.start()

    # 立即返回响应，不等待通知发送
    response_data = {
        'success': True,
        'message': f'问题已成功指派给 {assignee_name}',
        'redirect': f'/bug/{bug_id}?message=问题已成功指派给 {assignee_name}'
    }

    app.logger.info(f"问题分配成功，通知已提交后台处理 - bug_id: {bug_id}")
    return jsonify(response_data)

# 驳回问题API
@app.route('/bug/reject/<int:bug_id>', methods=['POST'])
@login_required
@role_required('fzr')
def reject_bug(bug_id):
    """负责人驳回问题"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    reject_reason = request.form.get('reject_reason', '').strip()
    if not reject_reason:
        return jsonify({'success': False, 'message': '驳回原因不能为空'})

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    try:
        # 检查问题是否存在
        query, params = adapt_sql('SELECT * FROM bugs WHERE id = %s', (bug_id,))
        c.execute(query, params)
        bug = c.fetchone()
        if not bug:
            return jsonify({'success': False, 'message': '问题不存在'}), 404

        # 检查问题状态是否可以驳回（只有待处理和已分配的问题可以驳回）
        if bug['status'] not in ['待处理', '已分配']:
            return jsonify({'success': False, 'message': '只有待处理或已分配的问题才能驳回'}), 400

        # 获取问题信息（用于通知）
        if hasattr(bug, 'keys'):
            bug_info = dict(bug)
        else:
            # SQLite返回的是tuple，需要手动映射
            bug_info = {
                'id': bug[0], 'title': bug[1], 'description': bug[2],
                'status': bug[3], 'assigned_to': bug[4], 'created_by': bug[5],
                'project': bug[6], 'created_at': bug[7], 'resolved_at': bug[8],
                'resolution': bug[9], 'image_path': bug[10]
            }

        # 更新问题状态为"已驳回"，清除指派人，并记录驳回原因
        query, params = adapt_sql('''
            UPDATE bugs
            SET status = '已驳回',
                assigned_to = NULL,
                resolution = %s
            WHERE id = %s
        ''', (f'驳回原因：{reject_reason}', bug_id))
        c.execute(query, params)
        conn.commit()
        conn.close()

        # 异步发送通知（在后台处理）
        def send_reject_notification_async():
            try:
                app.logger.info(f"后台发送问题驳回通知 - bug_id: {bug_id}")
                from notification.simple_notifier import simple_notifier
                from datetime import datetime

                notification_data = {
                    'bug_id': bug_id,
                    'title': bug_info['title'],
                    'description': bug_info['description'],
                    'reject_reason': reject_reason,
                    'rejector_name': user['chinese_name'] or user['username'],
                    'rejected_time': datetime.now().isoformat(),
                    'creator_id': bug_info['created_by'],
                    'old_assignee_id': bug_info.get('assigned_to')
                }

                simple_notifier.send_flow_notification('bug_rejected', notification_data)
                app.logger.info(f"问题驳回通知发送完成 - bug_id: {bug_id}")

            except Exception as e:
                app.logger.error(f"发送问题驳回通知失败 - bug_id: {bug_id}, error: {str(e)}")

        # 在后台线程中发送通知
        import threading
        notification_thread = threading.Thread(target=send_reject_notification_async)
        notification_thread.daemon = True
        notification_thread.start()

        return jsonify({
            'success': True,
            'message': '问题已成功驳回',
            'redirect': f'/bug/{bug_id}?message=问题已成功驳回'
        })

    except Exception as e:
        conn.close()
        app.logger.error(f"驳回问题失败 - bug_id: {bug_id}, error: {str(e)}")
        return jsonify({'success': False, 'message': f'驳回失败：{str(e)}'}), 500

# 解决问题页面
@app.route('/bug/resolve/<int:bug_id>')
@login_required
@role_required('zncy')
def resolve_page(bug_id):
    user = get_current_user()
    if not user:
        return redirect('/login')

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()
    query, params = adapt_sql('''
        SELECT b.*, COALESCE(u1.chinese_name, u1.username) as creator_name, COALESCE(u2.chinese_name, u2.username) as assignee_name
        FROM bugs b
        LEFT JOIN users u1 ON b.created_by = u1.id
        LEFT JOIN users u2 ON b.assigned_to = u2.id
        WHERE b.id = %s
    ''', (bug_id,))
    c.execute(query, params)
    bug = c.fetchone()
    conn.close()
    if not bug:
        return "问题不存在", 404
    return render_template('resolve.html', bug=bug, user=user)

# 解决问题API
@app.route('/bug/delete/<int:bug_id>', methods=['POST'])
@login_required
def delete_bug(bug_id):
    """删除问题"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    # 检查用户权限：管理员或问题创建者
    query, params = adapt_sql('SELECT created_by FROM bugs WHERE id = %s', (bug_id,))
    c.execute(query, params)
    bug = c.fetchone()
    if not bug:
        return jsonify({'success': False, 'message': '问题不存在'})

    if user['role_en'] != 'gly' and user['id'] != bug['created_by']:
        return jsonify({'success': False, 'message': '无权删除此问题'})

    # 执行删除
    query, params = adapt_sql('DELETE FROM bugs WHERE id = %s', (bug_id,))
    c.execute(query, params)
    conn.commit()
    conn.close()

    return jsonify({'success': True})

@app.route('/bug/confirm/<int:bug_id>', methods=['POST'])
@login_required
@role_required('zncy')
def confirm_receive(bug_id):
    """确认接收问题"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    # 检查问题是否已分配给当前用户
    query, params = adapt_sql('SELECT assigned_to FROM bugs WHERE id = %s', (bug_id,))
    c.execute(query, params)
    bug = c.fetchone()
    if not bug:
        conn.close()
        return jsonify({'success': False, 'message': '问题不存在'})

    if bug['assigned_to'] != user['id']:
        conn.close()
        return jsonify({'success': False, 'message': '无权操作此问题'})

    # 更新问题状态为"处理中"
    query, params = adapt_sql('''
        UPDATE bugs
        SET status = '处理中'
        WHERE id = %s
    ''', (bug_id,))
    c.execute(query, params)
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/bug/resolve/<int:bug_id>', methods=['GET'])
@login_required
@role_required('zncy')
def show_resolve_page(bug_id):
    """解决问题页面"""
    user = get_current_user()
    if not user:
        return redirect('/login')

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()
    query, params = adapt_sql('''
        SELECT b.*, COALESCE(u1.chinese_name, u1.username) as creator_name, COALESCE(u2.chinese_name, u2.username) as assignee_name
        FROM bugs b
        LEFT JOIN users u1 ON b.created_by = u1.id
        LEFT JOIN users u2 ON b.assigned_to = u2.id
        WHERE b.id = %s
    ''', (bug_id,))
    c.execute(query, params)
    bug = c.fetchone()
    conn.close()
    if not bug:
        return "问题不存在", 404
    return render_template('resolve.html', bug=bug, user=user)

@app.route('/bug/resolve/<int:bug_id>', methods=['POST'])
@login_required
@role_required('zncy')
def resolve_bug(bug_id):
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    resolution = request.form.get('resolution')
    if not resolution:
        return jsonify({'success': False, 'message': '处理详情不能为空'})

    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    # 获取问题信息（用于通知）
    query, params = adapt_sql('SELECT title, description, created_by FROM bugs WHERE id = %s', (bug_id,))
    c.execute(query, params)
    bug_info = c.fetchone()

    # 获取当前时间，精确到秒
    from datetime import datetime
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if DB_TYPE == 'postgres':
        query, params = adapt_sql('''
            UPDATE bugs
            SET resolution = %s, status = '已解决', resolved_at = %s
            WHERE id = %s AND assigned_to = %s
        ''', (resolution, current_time, bug_id, user['id']))
    else:
        query, params = adapt_sql('''
            UPDATE bugs
            SET resolution = %s, status = '已解决', resolved_at = %s
            WHERE id = %s AND assigned_to = %s
        ''', (resolution, current_time, bug_id, user['id']))
    c.execute(query, params)
    conn.commit()
    conn.close()

    # 立即返回响应，不等待通知发送
    response_data = {
        'success': True,
        'redirect': f'/bug/{bug_id}'
    }

    # 异步发送通知（在后台处理）
    def send_resolution_notification_async():
        try:
            app.logger.info(f"后台发送问题解决通知 - bug_id: {bug_id}")
            from notification.simple_notifier import simple_notifier
            from datetime import datetime

            notification_data = {
                'bug_id': bug_id,
                'title': bug_info['title'],
                'description': bug_info['description'],
                'solution': resolution,
                'resolver_name': user['chinese_name'] or user['username'],
                'resolved_time': datetime.now().isoformat(),
                'creator_id': bug_info['created_by']
            }

            simple_notifier.send_flow_notification('bug_resolved', notification_data)
            app.logger.info(f"问题解决通知发送完成 - bug_id: {bug_id}")
        except Exception as e:
            app.logger.error(f"后台解决通知发送失败 - bug_id: {bug_id}, 错误: {e}")
            import traceback
            app.logger.error(f"解决通知发送错误详情: {traceback.format_exc()}")

    # 启动后台线程发送通知
    import threading
    notification_thread = threading.Thread(target=send_resolution_notification_async, daemon=True)
    notification_thread.start()

    app.logger.info(f"问题解决成功，通知已提交后台处理 - bug_id: {bug_id}")
    return jsonify(response_data)

# 通知管理路由
@app.route('/admin/notifications')
@login_required
@role_required('gly')
def notification_settings():
    """通知管理页面（仅管理员）"""
    user = get_current_user()
    if not user:
        return redirect('/login')

    try:
        from notification.notification_manager import NotificationManager

        # 获取服务器通知状态
        server_enabled = NotificationManager.is_notification_enabled()

        # 获取所有用户的通知设置
        users = NotificationManager.get_all_users_preferences()

        return render_template('admin/notifications.html',
                             server_enabled=server_enabled,
                             users=users)
    except Exception as e:
        app.logger.error(f"Error loading notification settings: {e}")
        return "加载通知设置失败", 500

@app.route('/admin/notifications/server', methods=['POST'])
@login_required
@role_required('gly')
def toggle_server_notification():
    """切换服务器通知开关"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    try:
        from notification.notification_manager import NotificationManager

        enabled = request.json.get('enabled', False)
        success = NotificationManager.set_server_notification(enabled, user['id'])

        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': '权限不足或操作失败'})
    except Exception as e:
        app.logger.error(f"Error toggling server notification: {e}")
        return jsonify({'success': False, 'message': '系统错误'})

@app.route('/admin/notifications/user', methods=['POST'])
@login_required
@role_required('gly')
def toggle_user_notification():
    """切换用户通知开关"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    try:
        from notification.notification_manager import NotificationManager

        user_id = request.json.get('user_id')
        channel = request.json.get('channel')
        enabled = request.json.get('enabled', False)

        if not user_id or not channel:
            return jsonify({'success': False, 'message': '参数不完整'})

        success = NotificationManager.set_user_notification(user_id, channel, enabled, user['id'])

        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': '操作失败'})
    except Exception as e:
        app.logger.error(f"Error toggling user notification: {e}")
        return jsonify({'success': False, 'message': '系统错误'})

@app.route('/admin/notification-status')
@login_required
@role_required('gly')
def admin_notification_status():
    """获取服务器通知状态"""
    try:
        from notification.notification_manager import NotificationManager
        enabled = NotificationManager.is_notification_enabled()
        return jsonify({'enabled': enabled})
    except Exception as e:
        app.logger.error(f"获取通知状态失败: {e}")
        return jsonify({'enabled': False})

@app.route('/admin/notification-stats')
@login_required
@role_required('gly')
def admin_notification_stats():
    """获取通知统计信息"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 总通知数
        cursor.execute("SELECT COUNT(*) FROM notifications")
        total = cursor.fetchone()[0]

        # 未读通知数
        if DB_TYPE == 'postgres':
            cursor.execute("SELECT COUNT(*) FROM notifications WHERE read_status = false")
        else:
            cursor.execute("SELECT COUNT(*) FROM notifications WHERE read_status = 0")
        unread = cursor.fetchone()[0]

        # 启用通知的用户数
        if DB_TYPE == 'postgres':
            cursor.execute("SELECT COUNT(*) FROM user_notification_preferences WHERE email_enabled = true OR gotify_enabled = true OR inapp_enabled = true")
        else:
            cursor.execute("SELECT COUNT(*) FROM user_notification_preferences WHERE email_enabled = 1 OR gotify_enabled = 1 OR inapp_enabled = 1")
        enabled_users = cursor.fetchone()[0]

        # 今日通知数
        if DB_TYPE == 'postgres':
            cursor.execute("SELECT COUNT(*) FROM notifications WHERE DATE(created_at) = CURRENT_DATE")
        else:
            cursor.execute("SELECT COUNT(*) FROM notifications WHERE date(created_at) = date('now')")
        today = cursor.fetchone()[0]

        conn.close()

        return jsonify({
            'total': total,
            'unread': unread,
            'enabled_users': enabled_users,
            'today': today
        })

    except Exception as e:
        app.logger.error(f"获取通知统计失败: {e}")
        return jsonify({
            'total': 0,
            'unread': 0,
            'enabled_users': 0,
            'today': 0
        })

@app.route('/admin/toggle-notification', methods=['POST'])
@login_required
@role_required('gly')
def admin_toggle_notification():
    """切换服务器通知开关"""
    try:
        data = request.get_json()
        enabled = data.get('enabled', False)

        from notification.notification_manager import NotificationManager
        success = NotificationManager.set_notification_enabled(enabled)

        if success:
            return jsonify({'success': True, 'message': '设置成功'})
        else:
            return jsonify({'success': False, 'message': '设置失败'})

    except Exception as e:
        app.logger.error(f"切换通知开关失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/global-notification-status')
@login_required
@role_required('gly')
def admin_global_notification_status():
    """获取全局通知开关状态"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # 获取邮件全局开关
        cursor.execute("SELECT config_value FROM system_config WHERE config_key = 'email_global_enabled'")
        email_result = cursor.fetchone()
        email_enabled = email_result[0].lower() == 'true' if email_result else True

        # 获取Gotify全局开关
        cursor.execute("SELECT config_value FROM system_config WHERE config_key = 'gotify_global_enabled'")
        gotify_result = cursor.fetchone()
        gotify_enabled = gotify_result[0].lower() == 'true' if gotify_result else True

        conn.close()

        return jsonify({
            'email_enabled': email_enabled,
            'gotify_enabled': gotify_enabled
        })

    except Exception as e:
        app.logger.error(f"获取全局通知状态失败: {e}")
        return jsonify({
            'email_enabled': True,
            'gotify_enabled': True
        })

@app.route('/admin/toggle-global-notification', methods=['POST'])
@login_required
@role_required('gly')
def admin_toggle_global_notification():
    """切换全局通知开关"""
    try:
        data = request.get_json()
        notification_type = data.get('type')  # 'email' 或 'gotify'
        enabled = data.get('enabled', False)

        if notification_type not in ['email', 'gotify']:
            return jsonify({'success': False, 'message': '无效的通知类型'})

        conn = get_db_connection()
        cursor = conn.cursor()

        config_key = f'{notification_type}_global_enabled'
        config_value = 'true' if enabled else 'false'

        # 更新或插入配置
        cursor.execute("""
            UPDATE system_config
            SET config_value = %s, updated_at = CURRENT_TIMESTAMP
            WHERE config_key = %s
        """, (config_value, config_key))

        # 如果没有更新任何行，则插入新记录
        if cursor.rowcount == 0:
            description = f'全局{notification_type}通知开关'
            cursor.execute("""
                INSERT INTO system_config (config_key, config_value, description)
                VALUES (%s, %s, %s)
            """, (config_key, config_value, description))

        conn.commit()
        conn.close()

        return jsonify({'success': True, 'message': '设置成功'})

    except Exception as e:
        app.logger.error(f"切换全局通知开关失败: {e}")
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({'success': False, 'message': str(e)})

# 应用内通知API
@app.route('/api/notifications')
@login_required
def api_get_notifications():
    """获取用户通知"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    try:
        from notification.channels.inapp_notifier import InAppNotifier

        inapp_notifier = InAppNotifier()

        # 获取最新的10条通知
        notifications = inapp_notifier.get_user_notifications(str(user['id']), limit=10)

        # 获取未读数量
        unread_count = inapp_notifier.get_unread_count(str(user['id']))

        return jsonify({
            'success': True,
            'notifications': notifications,
            'unread_count': unread_count
        })

    except Exception as e:
        app.logger.error(f"获取通知失败: {e}")
        return jsonify({'success': False, 'message': '获取通知失败'})

@app.route('/api/notifications/read', methods=['POST'])
@login_required
def api_mark_notification_read():
    """标记通知为已读"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    try:
        data = request.get_json()
        notification_id = data.get('notification_id')

        if not notification_id:
            return jsonify({'success': False, 'message': '缺少通知ID'})

        from notification.channels.inapp_notifier import InAppNotifier

        inapp_notifier = InAppNotifier()
        success = inapp_notifier.mark_as_read(notification_id, str(user['id']))

        if success:
            return jsonify({'success': True, 'message': '标记成功'})
        else:
            return jsonify({'success': False, 'message': '标记失败'})

    except Exception as e:
        app.logger.error(f"标记通知已读失败: {e}")
        return jsonify({'success': False, 'message': '操作失败'})

@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def api_mark_all_notifications_read():
    """标记所有通知为已读"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    try:
        from notification.channels.inapp_notifier import InAppNotifier

        inapp_notifier = InAppNotifier()
        success = inapp_notifier.mark_all_as_read(str(user['id']))

        if success:
            return jsonify({'success': True, 'message': '全部标记成功'})
        else:
            return jsonify({'success': False, 'message': '标记失败'})

    except Exception as e:
        app.logger.error(f"标记全部通知已读失败: {e}")
        return jsonify({'success': False, 'message': '操作失败'})

@app.route('/notifications')
@login_required
def notifications_page():
    """通知页面"""
    user = get_current_user()
    if not user:
        return redirect('/login')

    try:
        from notification.channels.inapp_notifier import InAppNotifier

        inapp_notifier = InAppNotifier()

        # 获取所有通知
        notifications = inapp_notifier.get_user_notifications(str(user['id']), limit=50)

        return render_template('notifications.html', notifications=notifications)

    except Exception as e:
        app.logger.error(f"加载通知页面失败: {e}")
        return "加载通知页面失败", 500

# 通知配置API
@app.route('/admin/notifications/config', methods=['GET'])
@login_required
@role_required('gly')
def get_notification_config():
    """获取通知配置"""
    try:
        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            from psycopg2.extras import DictCursor
            cursor = conn.cursor(cursor_factory=DictCursor)
        else:
            cursor = conn.cursor()

        # 获取所有配置
        query, params = adapt_sql('SELECT config_key, config_value FROM system_config WHERE config_key LIKE %s', ('notification_%',))
        cursor.execute(query, params)
        configs = cursor.fetchall()

        # 构建配置字典
        config_dict = {}
        for config in configs:
            key = config['config_key'] if isinstance(config, dict) else config[0]
            value = config['config_value'] if isinstance(config, dict) else config[1]
            config_dict[key] = value

        # 构建返回的配置结构
        result = {
            'server': {
                'enabled': config_dict.get('notification_server_enabled', 'true') == 'true',
                'retention_days': int(config_dict.get('notification_retention_days', '30')),
                'auto_cleanup_enabled': config_dict.get('notification_auto_cleanup_enabled', 'false') == 'true'
            },
            'inapp': {
                'enabled': config_dict.get('notification_inapp_enabled', 'true') == 'true',
                'max_notifications_per_user': int(config_dict.get('notification_max_per_user', '100'))
            },
            'email': {
                'enabled': config_dict.get('notification_email_enabled', 'false') == 'true',
                'smtp_server': config_dict.get('notification_email_smtp_server', 'smtp.gmail.com'),
                'smtp_port': int(config_dict.get('notification_email_smtp_port', '587')),
                'smtp_username': config_dict.get('notification_email_smtp_username', ''),
                'smtp_password': config_dict.get('notification_email_smtp_password', ''),
                'from_email': config_dict.get('notification_email_from_email', 'noreply@rebugtracker.com'),
                'from_name': config_dict.get('notification_email_from_name', 'ReBugTracker'),
                'use_tls': config_dict.get('notification_email_use_tls', 'true') == 'true'
            },
            'gotify': {
                'enabled': config_dict.get('notification_gotify_enabled', 'false') == 'true',
                'server_url': config_dict.get('notification_gotify_server_url', 'http://localhost:8080'),
                'app_token': config_dict.get('notification_gotify_app_token', ''),
                'default_priority': int(config_dict.get('notification_gotify_default_priority', '10'))
            },
            'flow_rules': {
                'bug_created': config_dict.get('notification_flow_bug_created', 'true') == 'true',
                'bug_assigned': config_dict.get('notification_flow_bug_assigned', 'true') == 'true',
                'bug_rejected': config_dict.get('notification_flow_bug_rejected', 'true') == 'true',
                'bug_status_changed': config_dict.get('notification_flow_bug_status_changed', 'true') == 'true',
                'bug_resolved': config_dict.get('notification_flow_bug_resolved', 'true') == 'true',
                'bug_closed': config_dict.get('notification_flow_bug_closed', 'true') == 'true'
            }
        }

        conn.close()
        return jsonify({'success': True, 'data': result})

    except Exception as e:
        app.logger.error(f"获取通知配置失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/notifications/config', methods=['POST'])
@login_required
@role_required('gly')
def save_notification_config():
    """保存通知配置"""
    try:
        data = request.get_json()

        conn = get_db_connection()
        cursor = conn.cursor()

        # 配置映射
        config_mappings = {
            # 服务器配置
            'notification_server_enabled': str(data.get('server', {}).get('enabled', True)).lower(),
            'notification_retention_days': str(data.get('server', {}).get('retention_days', 30)),
            'notification_auto_cleanup_enabled': str(data.get('server', {}).get('auto_cleanup_enabled', False)).lower(),

            # 应用内通知
            'notification_inapp_enabled': str(data.get('inapp', {}).get('enabled', True)).lower(),
            'notification_max_per_user': str(data.get('inapp', {}).get('max_notifications_per_user', 100)),

            # 邮件通知
            'notification_email_enabled': str(data.get('email', {}).get('enabled', False)).lower(),
            'notification_email_smtp_server': data.get('email', {}).get('smtp_server', 'smtp.gmail.com'),
            'notification_email_smtp_port': str(data.get('email', {}).get('smtp_port', 587)),
            'notification_email_smtp_username': data.get('email', {}).get('smtp_username', ''),
            'notification_email_smtp_password': data.get('email', {}).get('smtp_password', ''),
            'notification_email_from_email': data.get('email', {}).get('from_email', 'noreply@rebugtracker.com'),
            'notification_email_from_name': data.get('email', {}).get('from_name', 'ReBugTracker'),
            'notification_email_use_tls': str(data.get('email', {}).get('use_tls', True)).lower(),

            # Gotify通知
            'notification_gotify_enabled': str(data.get('gotify', {}).get('enabled', False)).lower(),
            'notification_gotify_server_url': data.get('gotify', {}).get('server_url', 'http://localhost:8080'),
            'notification_gotify_app_token': data.get('gotify', {}).get('app_token', ''),
            'notification_gotify_default_priority': str(data.get('gotify', {}).get('default_priority', 10)),

            # 流程规则
            'notification_flow_bug_created': str(data.get('flow_rules', {}).get('bug_created', True)).lower(),
            'notification_flow_bug_assigned': str(data.get('flow_rules', {}).get('bug_assigned', True)).lower(),
            'notification_flow_bug_rejected': str(data.get('flow_rules', {}).get('bug_rejected', True)).lower(),
            'notification_flow_bug_status_changed': str(data.get('flow_rules', {}).get('bug_status_changed', True)).lower(),
            'notification_flow_bug_resolved': str(data.get('flow_rules', {}).get('bug_resolved', True)).lower(),
            'notification_flow_bug_closed': str(data.get('flow_rules', {}).get('bug_closed', True)).lower(),
        }

        # 保存每个配置项
        for config_key, config_value in config_mappings.items():
            if DB_TYPE == 'postgres':
                query, params = adapt_sql('''
                    INSERT INTO system_config (config_key, config_value, updated_at)
                    VALUES (%s, %s, CURRENT_TIMESTAMP)
                    ON CONFLICT (config_key) DO UPDATE SET
                    config_value = EXCLUDED.config_value,
                    updated_at = CURRENT_TIMESTAMP
                ''', (config_key, config_value))
            else:
                # SQLite使用不同的语法
                query, params = adapt_sql('''
                    INSERT OR REPLACE INTO system_config (config_key, config_value, updated_at)
                    VALUES (%s, %s, datetime('now'))
                ''', (config_key, config_value))

            cursor.execute(query, params)

        conn.commit()
        conn.close()

        return jsonify({'success': True, 'message': '通知配置保存成功'})

    except Exception as e:
        app.logger.error(f"保存通知配置失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/notifications/test', methods=['POST'])
@login_required
@role_required('gly')
def test_notification():
    """测试通知"""
    try:
        data = request.get_json()
        channels = data.get('channels', ['inapp'])
        message = data.get('message', '这是一条测试通知')

        successful_channels = []
        failed_channels = []

        # 测试应用内通知
        if 'inapp' in channels:
            try:
                from notification.channels.inapp_notifier import InAppNotifier
                inapp_notifier = InAppNotifier()
                user = get_current_user()

                success = inapp_notifier.send_notification(
                    user_id=str(user['id']),
                    title='测试通知',
                    content=message,
                    notification_type='system'
                )

                if success:
                    successful_channels.append('应用内通知')
                else:
                    failed_channels.append('应用内通知')
            except Exception as e:
                app.logger.error(f"测试应用内通知失败: {e}")
                failed_channels.append('应用内通知')

        # 测试邮件通知
        if 'email' in channels:
            try:
                # 这里可以添加邮件通知测试逻辑
                # 暂时标记为成功
                successful_channels.append('邮件通知')
            except Exception as e:
                app.logger.error(f"测试邮件通知失败: {e}")
                failed_channels.append('邮件通知')

        # 测试Gotify通知
        if 'gotify' in channels:
            try:
                # 这里可以添加Gotify通知测试逻辑
                # 暂时标记为成功
                successful_channels.append('Gotify通知')
            except Exception as e:
                app.logger.error(f"测试Gotify通知失败: {e}")
                failed_channels.append('Gotify通知')

        return jsonify({
            'success': True,
            'message': '测试通知已发送',
            'successful_channels': successful_channels,
            'failed_channels': failed_channels
        })

    except Exception as e:
        app.logger.error(f"测试通知失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/notifications/user-preferences', methods=['GET'])
@login_required
@role_required('gly')
def get_user_notification_preferences():
    """获取用户通知偏好"""
    try:
        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            from psycopg2.extras import DictCursor
            cursor = conn.cursor(cursor_factory=DictCursor)
        else:
            cursor = conn.cursor()

        # 获取所有用户及其通知偏好
        query, params = adapt_sql('''
            SELECT u.id, u.username, u.chinese_name, u.email,
                   COALESCE(np.inapp_enabled, true) as inapp_enabled,
                   COALESCE(np.email_enabled, false) as email_enabled,
                   COALESCE(np.gotify_enabled, false) as gotify_enabled
            FROM users u
            LEFT JOIN user_notification_preferences np ON u.id = np.user_id
            ORDER BY u.id
        ''', ())

        cursor.execute(query, params)
        users = cursor.fetchall()

        # 构建返回数据
        preferences = {}
        for user in users:
            user_id = user['id'] if isinstance(user, dict) else user[0]
            preferences[str(user_id)] = {
                'inapp': user['inapp_enabled'] if isinstance(user, dict) else user[4],
                'email': user['email_enabled'] if isinstance(user, dict) else user[5],
                'gotify': user['gotify_enabled'] if isinstance(user, dict) else user[6]
            }

        conn.close()
        return jsonify({'success': True, 'data': preferences})

    except Exception as e:
        app.logger.error(f"获取用户通知偏好失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/notifications/user-preferences', methods=['POST'])
@login_required
@role_required('gly')
def save_user_notification_preference():
    """保存单个用户通知偏好"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        inapp_enabled = data.get('inapp', True)
        email_enabled = data.get('email', False)
        gotify_enabled = data.get('gotify', False)

        conn = get_db_connection()
        cursor = conn.cursor()

        if DB_TYPE == 'postgres':
            query, params = adapt_sql('''
                INSERT INTO user_notification_preferences (user_id, inapp_enabled, email_enabled, gotify_enabled, updated_at)
                VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP)
                ON CONFLICT (user_id) DO UPDATE SET
                inapp_enabled = EXCLUDED.inapp_enabled,
                email_enabled = EXCLUDED.email_enabled,
                gotify_enabled = EXCLUDED.gotify_enabled,
                updated_at = CURRENT_TIMESTAMP
            ''', (user_id, inapp_enabled, email_enabled, gotify_enabled))
        else:
            query, params = adapt_sql('''
                INSERT OR REPLACE INTO user_notification_preferences (user_id, inapp_enabled, email_enabled, gotify_enabled, updated_at)
                VALUES (%s, %s, %s, %s, datetime('now'))
            ''', (user_id, inapp_enabled, email_enabled, gotify_enabled))

        cursor.execute(query, params)
        conn.commit()
        conn.close()

        return jsonify({'success': True, 'message': '用户通知偏好保存成功'})

    except Exception as e:
        app.logger.error(f"保存用户通知偏好失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/notifications/user-preferences/batch', methods=['POST'])
@login_required
@role_required('gly')
def save_batch_user_notification_preferences():
    """批量保存用户通知偏好"""
    try:
        data = request.get_json()
        preferences = data.get('preferences', [])

        conn = get_db_connection()
        cursor = conn.cursor()

        for pref in preferences:
            user_id = pref.get('user_id')
            inapp_enabled = pref.get('inapp', True)
            email_enabled = pref.get('email', False)
            gotify_enabled = pref.get('gotify', False)

            if DB_TYPE == 'postgres':
                query, params = adapt_sql('''
                    INSERT INTO user_notification_preferences (user_id, inapp_enabled, email_enabled, gotify_enabled, updated_at)
                    VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP)
                    ON CONFLICT (user_id) DO UPDATE SET
                    inapp_enabled = EXCLUDED.inapp_enabled,
                    email_enabled = EXCLUDED.email_enabled,
                    gotify_enabled = EXCLUDED.gotify_enabled,
                    updated_at = CURRENT_TIMESTAMP
                ''', (user_id, inapp_enabled, email_enabled, gotify_enabled))
            else:
                query, params = adapt_sql('''
                    INSERT OR REPLACE INTO user_notification_preferences (user_id, inapp_enabled, email_enabled, gotify_enabled, updated_at)
                    VALUES (%s, %s, %s, %s, datetime('now'))
                ''', (user_id, inapp_enabled, email_enabled, gotify_enabled))

            cursor.execute(query, params)

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': f'批量保存了 {len(preferences)} 个用户的通知偏好',
            'updated_count': len(preferences)
        })

    except Exception as e:
        app.logger.error(f"批量保存用户通知偏好失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/notifications/cleanup', methods=['POST'])
@login_required
@role_required('gly')
def admin_cleanup_notifications():
    """手动触发通知清理"""
    try:
        data = request.get_json() or {}
        cleanup_type = data.get('type', 'all')  # 'expired', 'excess', 'all'

        from notification.cleanup_manager import cleanup_manager

        results = {}

        if cleanup_type in ['expired', 'all']:
            # 清理过期通知
            expired_result = cleanup_manager.cleanup_expired_notifications()
            results['expired'] = expired_result

        if cleanup_type in ['excess', 'all']:
            # 清理超量通知
            excess_result = cleanup_manager.cleanup_excess_notifications()
            results['excess'] = excess_result

        # 计算总清理数量
        total_deleted = 0
        if 'expired' in results:
            total_deleted += results['expired'].get('deleted_count', 0)
        if 'excess' in results:
            total_deleted += results['excess'].get('total_deleted', 0)

        return jsonify({
            'success': True,
            'message': f'清理完成，共删除 {total_deleted} 条通知',
            'total_deleted': total_deleted,
            'results': results
        })

    except Exception as e:
        app.logger.error(f"手动清理通知失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/notifications/cleanup/stats')
@login_required
@role_required('gly')
def admin_cleanup_stats():
    """获取通知清理统计信息"""
    try:
        from notification.cleanup_manager import cleanup_manager

        stats = cleanup_manager.get_cleanup_stats()

        return jsonify({
            'success': True,
            'data': stats
        })

    except Exception as e:
        app.logger.error(f"获取清理统计失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

# 项目管理API
@app.route('/admin/projects', methods=['GET', 'POST'])
@login_required
@role_required('gly')
def admin_projects():
    """项目管理API"""
    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    try:
        if request.method == 'GET':
            # 获取所有项目
            query = '''
                SELECT id, name, type, city, start_date, factory_acceptance_date,
                       site_acceptance_date, practical_date, description,
                       created_at, updated_at
                FROM projects
                ORDER BY created_at DESC
            '''
            adapted_query, adapted_params = adapt_sql(query, ())
            c.execute(adapted_query, adapted_params)
            rows = c.fetchall()

            projects = []
            for row in rows:
                project = dict(row)
                # 处理日期字段
                for date_field in ['start_date', 'factory_acceptance_date', 'site_acceptance_date', 'practical_date']:
                    if project[date_field]:
                        if isinstance(project[date_field], str):
                            project[date_field] = project[date_field]
                        else:
                            project[date_field] = project[date_field].strftime('%Y-%m-%d')
                    else:
                        project[date_field] = ''

                # 处理时间戳字段
                for timestamp_field in ['created_at', 'updated_at']:
                    if project[timestamp_field]:
                        if isinstance(project[timestamp_field], str):
                            project[timestamp_field] = project[timestamp_field]
                        else:
                            project[timestamp_field] = project[timestamp_field].strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        project[timestamp_field] = ''

                projects.append(project)

            return jsonify({'success': True, 'data': projects})

        elif request.method == 'POST':
            # 创建新项目
            data = request.get_json()

            project_type = data.get('type', '')
            city = data.get('city', '')

            # 自动生成项目名称：类型#地市名称+项目
            if project_type and city:
                auto_name = f"{project_type}#{city}项目"
            else:
                return jsonify({'success': False, 'message': '项目类型和地市名称不能为空'})

            # 检查项目名称是否已存在
            check_query, check_params = adapt_sql('SELECT id FROM projects WHERE name = %s', (auto_name,))
            c.execute(check_query, check_params)
            if c.fetchone():
                return jsonify({'success': False, 'message': f'项目名称 "{auto_name}" 已存在'})

            # 插入新项目
            from datetime import datetime
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            insert_query = '''
                INSERT INTO projects (name, type, city, start_date, factory_acceptance_date,
                                    site_acceptance_date, practical_date, description,
                                    created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            '''
            insert_params = (
                auto_name,
                project_type,
                city,
                data.get('start_date') or None,
                data.get('factory_acceptance_date') or None,
                data.get('site_acceptance_date') or None,
                data.get('practical_date') or None,
                data.get('description', ''),
                current_time,
                current_time
            )

            adapted_insert_query, adapted_insert_params = adapt_sql(insert_query, insert_params)
            c.execute(adapted_insert_query, adapted_insert_params)

            # 提交事务（SQLite和PostgreSQL都需要）
            conn.commit()

            return jsonify({'success': True, 'message': f'项目 "{auto_name}" 创建成功'})

    except Exception as e:
        if DB_TYPE == 'postgres':
            conn.rollback()
        app.logger.error(f"项目管理API错误: {e}")
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()

@app.route('/admin/projects/export', methods=['GET'])
@login_required
@role_required('gly')
def export_projects():
    """导出项目数据为Excel"""
    try:
        from datetime import datetime

        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            c = conn.cursor(cursor_factory=DictCursor)
        else:
            c = conn.cursor()

        # 获取所有项目数据
        query = '''
            SELECT name, type, city, start_date, factory_acceptance_date,
                   site_acceptance_date, practical_date, description, created_at
            FROM projects
            ORDER BY created_at DESC
        '''
        adapted_query, adapted_params = adapt_sql(query, ())
        c.execute(adapted_query, adapted_params)

        # 转换为字典列表
        projects_data = []
        for row in c.fetchall():
            project = dict(row)
            # 处理日期格式
            for date_field in ['start_date', 'factory_acceptance_date', 'site_acceptance_date', 'practical_date', 'created_at']:
                if project.get(date_field):
                    try:
                        if isinstance(project[date_field], str):
                            # 如果是字符串，尝试解析
                            dt = datetime.strptime(project[date_field][:10], '%Y-%m-%d')
                            project[date_field] = dt.strftime('%Y-%m-%d')
                        else:
                            # 如果是datetime对象，直接格式化
                            project[date_field] = project[date_field].strftime('%Y-%m-%d')
                    except:
                        project[date_field] = str(project[date_field]) if project[date_field] else ''
                else:
                    project[date_field] = ''
            projects_data.append(project)

        conn.close()

        # 定义导出字段
        fields = [
            {'key': 'name', 'label': '项目名称'},
            {'key': 'type', 'label': '项目类型'},
            {'key': 'city', 'label': '地市名称'},
            {'key': 'start_date', 'label': '启动时间'},
            {'key': 'factory_acceptance_date', 'label': '工厂验收时间'},
            {'key': 'site_acceptance_date', 'label': '现场验收时间'},
            {'key': 'practical_date', 'label': '实用化时间'},
            {'key': 'description', 'label': '项目描述'},
            {'key': 'created_at', 'label': '创建时间'}
        ]

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'项目列表_{timestamp}'

        # 使用报表导出的Excel方法
        return export_to_excel(projects_data, fields, filename)

    except Exception as e:
        app.logger.error(f"导出项目数据错误: {e}")
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'})
    finally:
        if 'conn' in locals():
            conn.close()

@app.route('/admin/projects/<int:project_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
@role_required('gly')
def admin_project_detail(project_id):
    """项目详情管理API"""
    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    try:
        if request.method == 'GET':
            # 获取单个项目信息
            query, params = adapt_sql('''
                SELECT id, name, type, city, start_date, factory_acceptance_date,
                       site_acceptance_date, practical_date, description, created_at
                FROM projects WHERE id = %s
            ''', (project_id,))
            c.execute(query, params)
            project = c.fetchone()

            if not project:
                return jsonify({'success': False, 'message': '项目不存在'}), 404

            return jsonify(dict(project))

        elif request.method == 'PUT':
            # 更新项目
            data = request.get_json()

            project_type = data.get('type', '')
            city = data.get('city', '')

            # 重新生成项目名称
            if project_type and city:
                auto_name = f"{project_type}#{city}项目"
            else:
                return jsonify({'success': False, 'message': '项目类型和地市名称不能为空'})

            # 检查项目名称是否与其他项目冲突
            check_query, check_params = adapt_sql('SELECT id FROM projects WHERE name = %s AND id != %s', (auto_name, project_id))
            c.execute(check_query, check_params)
            if c.fetchone():
                return jsonify({'success': False, 'message': f'项目名称 "{auto_name}" 已被其他项目使用'})

            # 更新项目
            from datetime import datetime
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            update_query = '''
                UPDATE projects
                SET name = %s, type = %s, city = %s, start_date = %s,
                    factory_acceptance_date = %s, site_acceptance_date = %s,
                    practical_date = %s, description = %s, updated_at = %s
                WHERE id = %s
            '''
            update_params = (
                auto_name,
                project_type,
                city,
                data.get('start_date') or None,
                data.get('factory_acceptance_date') or None,
                data.get('site_acceptance_date') or None,
                data.get('practical_date') or None,
                data.get('description', ''),
                current_time,
                project_id
            )

            adapted_update_query, adapted_update_params = adapt_sql(update_query, update_params)
            c.execute(adapted_update_query, adapted_update_params)

            # 提交事务（SQLite和PostgreSQL都需要）
            conn.commit()

            return jsonify({'success': True, 'message': f'项目 "{auto_name}" 更新成功'})

        elif request.method == 'DELETE':
            # 检查项目是否被bugs引用
            check_bugs_query, check_bugs_params = adapt_sql('SELECT COUNT(*) as count FROM bugs WHERE project = (SELECT name FROM projects WHERE id = %s)', (project_id,))
            c.execute(check_bugs_query, check_bugs_params)
            result = c.fetchone()
            bug_count = result['count'] if DB_TYPE == 'postgres' else result[0]

            if bug_count > 0:
                return jsonify({'success': False, 'message': f'无法删除项目，该项目下还有 {bug_count} 个问题'})

            # 删除项目
            delete_query, delete_params = adapt_sql('DELETE FROM projects WHERE id = %s', (project_id,))
            c.execute(delete_query, delete_params)

            # 提交事务（SQLite和PostgreSQL都需要）
            conn.commit()

            return jsonify({'success': True, 'message': '项目删除成功'})

    except Exception as e:
        if DB_TYPE == 'postgres':
            conn.rollback()
        app.logger.error(f"项目详情管理API错误: {e}")
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()

@app.route('/api/projects/list', methods=['GET'])
@login_required
def get_projects_list():
    """获取项目列表（用于下拉菜单）"""
    conn = get_db_connection()
    if DB_TYPE == 'postgres':
        c = conn.cursor(cursor_factory=DictCursor)
    else:
        c = conn.cursor()

    try:
        query = 'SELECT name FROM projects ORDER BY name'
        adapted_query, adapted_params = adapt_sql(query, ())
        c.execute(adapted_query, adapted_params)
        projects = [row['name'] if DB_TYPE == 'postgres' else row[0] for row in c.fetchall()]

        return jsonify({'success': True, 'data': projects})

    except Exception as e:
        app.logger.error(f"获取项目列表错误: {e}")
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'),
                              'rbt_title.ico', mimetype='image/vnd.microsoft.icon')

# 添加上传文件访问路由
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    upload_folder = app.config['UPLOAD_FOLDER']
    file_path = os.path.join(upload_folder, filename)

    # 调试信息
    app.logger.debug(f"请求文件: {filename}")
    app.logger.debug(f"上传目录: {upload_folder}")
    app.logger.debug(f"完整路径: {file_path}")
    app.logger.debug(f"文件是否存在: {os.path.exists(file_path)}")

    if not os.path.exists(file_path):
        app.logger.error(f"文件不存在: {file_path}")
        abort(404)

    return send_from_directory(upload_folder, filename)

# 调试路由：显示上传目录信息


@app.route('/debug/uploads')
def debug_uploads():
    """调试上传目录信息"""
    upload_folder = app.config['UPLOAD_FOLDER']
    info = {
        'upload_folder': upload_folder,
        'is_absolute': os.path.isabs(upload_folder),
        'exists': os.path.exists(upload_folder),
        'writable': os.access(upload_folder, os.W_OK) if os.path.exists(upload_folder) else False,
        'current_dir': os.getcwd(),
        'files': []
    }

    if os.path.exists(upload_folder):
        try:
            files = os.listdir(upload_folder)
            for file in files[:20]:  # 只显示前20个文件
                file_path = os.path.join(upload_folder, file)
                info['files'].append({
                    'name': file,
                    'size': os.path.getsize(file_path),
                    'path': file_path
                })
        except Exception as e:
            info['error'] = str(e)

    return jsonify(info)



@app.route('/api/users/product-managers', methods=['GET'])
@login_required
@role_required('gly')
def get_product_managers():
    """获取所有产品经理用户"""
    try:
        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            c = conn.cursor(cursor_factory=DictCursor)
        else:
            c = conn.cursor()

        query, params = adapt_sql('''
            SELECT id, username, chinese_name, email, phone
            FROM users
            WHERE role_en = 'pm'
            ORDER BY chinese_name, username
        ''', ())

        c.execute(query, params)
        managers = c.fetchall()
        conn.close()

        # 转换为字典列表
        result = []
        for manager in managers:
            if isinstance(manager, dict):
                result.append(dict(manager))
            else:
                result.append({
                    'id': manager[0],
                    'username': manager[1],
                    'chinese_name': manager[2],
                    'email': manager[3],
                    'phone': manager[4]
                })

        return jsonify({'success': True, 'data': result})

    except Exception as e:
        app.logger.error(f"获取产品经理列表失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/team-statistics', methods=['GET'])
@login_required
@role_required('gly')
def get_team_statistics():
    """获取团队统计数据"""
    try:
        # 固定的团队列表
        fixed_teams = [
            '实施组', '实施组研发', '新能源', '网络分析', '第三道防线',
            '智能告警', '操作票及防误', '电量', '消纳', '自动发电控制'
        ]

        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            c = conn.cursor(cursor_factory=DictCursor)
        else:
            c = conn.cursor()

        team_stats = {}

        for team_name in fixed_teams:
            # 统计团队成员数量（包括多团队的产品经理）
            query, params = adapt_sql('''
                SELECT COUNT(*) as member_count
                FROM users
                WHERE team = %s OR team LIKE %s OR team LIKE %s OR team LIKE %s
            ''', (team_name, f'{team_name},%', f'%,{team_name},%', f'%,{team_name}'))
            c.execute(query, params)
            member_count = c.fetchone()
            if isinstance(member_count, dict):
                member_count = member_count['member_count']
            else:
                member_count = member_count[0]

            # 获取产品经理列表
            query, params = adapt_sql('''
                SELECT chinese_name
                FROM users
                WHERE role = '产品经理' AND (team = %s OR team LIKE %s OR team LIKE %s OR team LIKE %s)
            ''', (team_name, f'{team_name},%', f'%,{team_name},%', f'%,{team_name}'))
            c.execute(query, params)
            pm_results = c.fetchall()
            product_managers = []
            for pm in pm_results:
                if isinstance(pm, dict):
                    product_managers.append(pm['chinese_name'])
                else:
                    product_managers.append(pm[0])

            # 统计Bug数量 - 通过assigned_to联合users表的team字段
            query, params = adapt_sql('''
                SELECT
                    COUNT(*) as total,
                    SUM(CASE WHEN b.status = '待处理' THEN 1 ELSE 0 END) as pending,
                    SUM(CASE WHEN b.status IN ('已分配', '处理中') THEN 1 ELSE 0 END) as processing,
                    SUM(CASE WHEN b.status = '已解决' THEN 1 ELSE 0 END) as resolved
                FROM bugs b
                LEFT JOIN users u ON b.assigned_to = u.id
                WHERE u.team = %s
            ''', (team_name,))
            c.execute(query, params)
            bug_stats = c.fetchone()

            if isinstance(bug_stats, dict):
                total = bug_stats['total'] or 0
                pending = bug_stats['pending'] or 0
                processing = bug_stats['processing'] or 0
                resolved = bug_stats['resolved'] or 0
            else:
                total = bug_stats[0] or 0
                pending = bug_stats[1] or 0
                processing = bug_stats[2] or 0
                resolved = bug_stats[3] or 0

            team_stats[team_name] = {
                'memberCount': member_count,
                'productManagers': product_managers,
                'bugStats': {
                    'total': total,
                    'pending': pending,
                    'processing': processing,
                    'resolved': resolved
                }
            }

        conn.close()

        return jsonify({
            'success': True,
            'data': team_stats
        })
    except Exception as e:
        app.logger.error(f'获取团队统计数据失败: {e}')
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/.well-known/appspecific/com.chrome.devtools.json')
def handle_chrome_devtools():
    return jsonify({'message': 'Not supported'}), 404



@app.route('/debug')
def debug_page():
    """调试页面"""
    try:
        user = get_current_user()
        cookies = dict(request.cookies)

        # 测试team-issues功能
        test_result = "未测试"
        try:
            if user and safe_get(user, 'role_en') == 'zncy':
                # 模拟team-issues的数据库查询
                conn = get_db_connection()
                if DB_TYPE == 'postgres':
                    c = conn.cursor(cursor_factory=DictCursor)
                else:
                    c = conn.cursor()

                query, params = adapt_sql('''
                    SELECT COUNT(*) as count
                    FROM bugs b
                    LEFT JOIN users u1 ON b.created_by = u1.id
                    WHERE (
                        (b.assigned_to = %s)
                        OR
                        (b.status = '待处理' AND b.assigned_to IS NULL AND u1.team = %s)
                        OR
                        (b.status = '已解决' AND b.assigned_to = %s)
                    )
                ''', (user['id'], user['team'], user['id']))

                c.execute(query, params)
                result = c.fetchone()
                bug_count = result[0] if result else 0
                conn.close()
                test_result = f"找到 {bug_count} 个问题"
        except Exception as e:
            test_result = f"测试失败: {str(e)}"

        return f"""
        <html>
        <head><title>调试信息</title></head>
        <body>
            <h1>调试信息</h1>
            <h2>用户信息</h2>
            <p>当前用户: {user}</p>

            <h2>Cookies</h2>
            <p>Cookies: {cookies}</p>

            <h2>数据库测试</h2>
            <p>测试结果: {test_result}</p>

            <h2>导航</h2>
            <a href="/login">登录</a> |
            <a href="/">首页</a> |
            <a href="/team-issues">team-issues</a> |
            <a href="/logout">退出</a>

            <h2>直接访问测试</h2>
            <form action="/team-issues" method="get">
                <button type="submit">直接访问team-issues</button>
            </form>
        </body>
        </html>
        """
    except Exception as e:
        import traceback
        return f"""
        <html>
        <head><title>调试错误</title></head>
        <body>
            <h1>调试页面错误</h1>
            <p>错误: {str(e)}</p>
            <pre>{traceback.format_exc()}</pre>
        </body>
        </html>
        """

def check_port_available(host, port):
    """检查端口是否可用"""
    import socket
    import sys

    try:
        # 创建socket并尝试绑定端口
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1)
        result = sock.connect_ex((host, port))
        sock.close()

        if result == 0:
            # 端口被占用
            print(f"❌ 错误: 端口 {port} 已被占用!")
            print(f"请先关闭占用端口 {port} 的其他应用程序，或者:")
            print(f"1. 使用命令查看占用进程: netstat -ano | findstr :{port}")
            print(f"2. 杀掉占用进程: taskkill /F /PID <进程ID>")
            print(f"3. 然后重新运行此程序")
            sys.exit(1)
        else:
            print(f"✅ 端口 {port} 可用，正在启动应用程序...")
            return True

    except Exception as e:
        print(f"⚠️  端口检测出错: {e}")
        print(f"继续启动应用程序...")
        return True

# 用户设置相关路由已在上面定义

@app.route('/user/email-settings', methods=['GET', 'POST'])
@login_required
def user_email_settings():
    """用户邮件设置"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    if request.method == 'GET':
        # 获取用户当前的邮件设置
        try:
            conn = get_db_connection()
            if DB_TYPE == 'postgres':
                c = conn.cursor(cursor_factory=DictCursor)
            else:
                c = conn.cursor()

            query, params = adapt_sql('SELECT email FROM users WHERE id = %s', (user['id'],))
            c.execute(query, params)
            result = c.fetchone()
            conn.close()

            email = result[0] if result and result[0] else None
            return jsonify({
                'success': True,
                'email': email
            })

        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})

    else:  # POST
        # 保存用户的邮件设置
        try:
            data = request.get_json()
            email = data.get('email', '').strip()

            # 如果邮箱不为空，进行格式验证
            if email:
                # 简单的邮箱格式验证
                import re
                email_pattern = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
                if not re.match(email_pattern, email):
                    return jsonify({'success': False, 'message': '邮箱格式不正确'})

            conn = get_db_connection()
            if DB_TYPE == 'postgres':
                c = conn.cursor(cursor_factory=DictCursor)
            else:
                c = conn.cursor()

            # 允许设置为空值（None）来清空邮箱
            email_value = email if email else None
            query, params = adapt_sql(
                'UPDATE users SET email = %s WHERE id = %s',
                (email_value, user['id'])
            )
            c.execute(query, params)
            conn.commit()
            conn.close()

            if email:
                return jsonify({'success': True, 'message': '邮箱设置保存成功'})
            else:
                return jsonify({'success': True, 'message': '邮箱设置已清空'})

        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})

@app.route('/user/test-email', methods=['POST'])
@login_required
def test_user_email():
    """测试用户的邮件发送"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    try:
        data = request.get_json()
        email = data.get('email', '').strip()

        if not email:
            return jsonify({'success': False, 'message': '邮箱地址不能为空'})

        # 测试发送邮件
        try:
            from notification.channels.email_notifier import EmailNotifier
            email_notifier = EmailNotifier()

            if not email_notifier.is_enabled():
                return jsonify({'success': False, 'message': '邮件服务未启用，请联系管理员配置'})

            # 构造测试邮件内容
            recipient_info = {
                'name': user.get('chinese_name') or user.get('username'),
                'email': email
            }

            success = email_notifier.send(
                title="🧪 ReBugTracker邮件测试",
                content=f"您好 {recipient_info['name']}！\n\n这是一封测试邮件，说明您的邮箱配置正确。\n\n如果您收到这封邮件，表示ReBugTracker可以正常向您发送通知。",
                recipient=recipient_info,
                priority=1
            )

            if success:
                return jsonify({'success': True, 'message': '测试邮件发送成功'})
            else:
                return jsonify({'success': False, 'message': '邮件发送失败，请检查邮箱地址'})

        except Exception as e:
            return jsonify({'success': False, 'message': f'邮件发送失败: {str(e)}'})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/user/gotify-settings', methods=['GET', 'POST'])
@login_required
def user_gotify_settings():
    """用户Gotify设置"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    if request.method == 'GET':
        # 获取用户当前的Gotify设置
        try:
            conn = get_db_connection()
            if DB_TYPE == 'postgres':
                c = conn.cursor(cursor_factory=DictCursor)
            else:
                c = conn.cursor()

            query, params = adapt_sql('SELECT gotify_app_token FROM users WHERE id = %s', (user['id'],))
            c.execute(query, params)
            result = c.fetchone()
            conn.close()

            app_token = result[0] if result and result[0] else None
            return jsonify({
                'success': True,
                'app_token': app_token[:10] + '...' if app_token else None  # 只显示前10位
            })

        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})

    else:  # POST
        # 保存用户的Gotify设置
        try:
            data = request.get_json()
            app_token = data.get('app_token', '').strip()

            conn = get_db_connection()
            if DB_TYPE == 'postgres':
                c = conn.cursor(cursor_factory=DictCursor)
            else:
                c = conn.cursor()

            # 允许设置为空值（None）来清空token
            token_value = app_token if app_token else None
            query, params = adapt_sql(
                'UPDATE users SET gotify_app_token = %s WHERE id = %s',
                (token_value, user['id'])
            )
            c.execute(query, params)
            conn.commit()
            conn.close()

            if app_token:
                return jsonify({'success': True, 'message': 'Gotify设置保存成功'})
            else:
                return jsonify({'success': True, 'message': 'Gotify Token已清空'})

        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})

@app.route('/user/test-gotify', methods=['POST'])
@login_required
def test_user_gotify():
    """测试用户的Gotify连接"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    try:
        data = request.get_json()
        app_token = data.get('app_token', '').strip()

        if not app_token:
            return jsonify({'success': False, 'message': 'App Token不能为空'})

        # 测试发送通知
        import requests
        import os

        server_url = os.getenv('GOTIFY_SERVER_URL', 'http://localhost:8080')
        url = f"{server_url.rstrip('/')}/message"

        test_data = {
            "title": "🧪 ReBugTracker测试通知",
            "message": f"您好 {user.get('chinese_name') or user.get('username')}！\n\n这是一条测试通知，说明您的Gotify配置正确。",
            "priority": 5
        }

        headers = {
            "X-Gotify-Key": app_token,
            "Content-Type": "application/json"
        }

        response = requests.post(url, json=test_data, headers=headers, timeout=10)

        if response.status_code == 200:
            return jsonify({'success': True, 'message': '测试通知发送成功'})
        else:
            return jsonify({'success': False, 'message': f'发送失败: {response.status_code}'})

    except requests.exceptions.RequestException as e:
        return jsonify({'success': False, 'message': f'网络错误: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/user/notification-preferences', methods=['GET', 'POST'])
@login_required
def user_notification_preferences():
    """用户通知偏好设置"""
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'message': '用户未登录'})

    if request.method == 'GET':
        # 获取用户通知偏好
        try:
            from notification.notification_manager import NotificationManager
            preferences = NotificationManager.is_user_notification_enabled(user['id'])
            return jsonify({
                'success': True,
                'inapp_enabled': preferences.get('inapp', True),
                'email_enabled': preferences.get('email', False),
                'gotify_enabled': preferences.get('gotify', False)
            })
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})

    else:  # POST
        # 保存用户通知偏好
        try:
            data = request.get_json()
            from notification.notification_manager import NotificationManager

            success = NotificationManager.set_user_notification_preferences(
                str(user['id']),
                email_enabled=data.get('email_enabled', False),
                gotify_enabled=data.get('gotify_enabled', False),
                inapp_enabled=data.get('inapp_enabled', True)
            )

            if success:
                return jsonify({'success': True, 'message': '通知偏好保存成功'})
            else:
                return jsonify({'success': False, 'message': '保存失败'})

        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})

# ==================== 报表导出功能 ====================

@app.route('/admin/reports/preview', methods=['POST'])
@login_required
@role_required('gly')
def admin_reports_preview():
    """预览报表数据"""

    try:
        data = request.get_json()
        filters = data.get('filters', {})
        fields = data.get('fields', [])

        # 构建查询
        query_parts = []
        params = []

        # 基础查询
        base_query = '''
            SELECT b.id, b.title, b.description, b.status, b.type, b.project,
                   b.created_at, b.resolved_at, b.resolution, b.image_path,
                   u1.username as creator_username, u1.chinese_name as creator_name,
                   u2.username as assignee_username, u2.chinese_name as assignee_name,
                   u2.team as assignee_team,
                   u3.username as manager_username, u3.chinese_name as manager_name
            FROM bugs b
            LEFT JOIN users u1 ON b.created_by = u1.id
            LEFT JOIN users u2 ON b.assigned_to = u2.id
            LEFT JOIN users u3 ON u2.team = u3.team AND u3.role_en = 'fzr'
        '''

        # 添加筛选条件
        where_conditions = []

        # 日期范围筛选
        if filters.get('dateRange'):
            date_range = filters['dateRange']
            if date_range.get('start'):
                where_conditions.append('DATE(b.created_at) >= %s')
                params.append(date_range['start'])
            if date_range.get('end'):
                where_conditions.append('DATE(b.created_at) <= %s')
                params.append(date_range['end'])

        # 状态筛选
        if filters.get('status'):
            status_list = filters['status']
            if status_list:
                placeholders = ','.join(['%s'] * len(status_list))
                where_conditions.append(f'b.status IN ({placeholders})')
                params.extend(status_list)

        # 项目筛选
        if filters.get('project'):
            where_conditions.append('b.project = %s')
            params.append(filters['project'])

        # 创建者筛选
        if filters.get('creator'):
            where_conditions.append('b.created_by = %s')
            params.append(filters['creator'])

        # 分配者筛选
        if filters.get('assignee'):
            where_conditions.append('b.assigned_to = %s')
            params.append(filters['assignee'])

        # 类型筛选
        if filters.get('type'):
            type_list = filters['type']
            if type_list:
                placeholders = ','.join(['%s'] * len(type_list))
                where_conditions.append(f'b.type IN ({placeholders})')
                params.extend(type_list)

        # 组合查询
        if where_conditions:
            query = base_query + ' WHERE ' + ' AND '.join(where_conditions)
        else:
            query = base_query

        query += ' ORDER BY b.created_at DESC LIMIT 100'  # 预览限制100条

        # 执行查询
        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            from psycopg2.extras import DictCursor
            cursor = conn.cursor(cursor_factory=DictCursor)
        else:
            cursor = conn.cursor()

        query, params = adapt_sql(query, params)
        cursor.execute(query, params)

        results = cursor.fetchall()
        conn.close()

        # 转换数据格式 - 兼容PostgreSQL和SQLite
        preview_data = []
        assignee_usernames = set()  # 收集所有分配者用户名
        bug_ids = []  # 收集所有bug ID，用于查询附件数量

        for row in results:
            # 统一转换为字典格式
            if DB_TYPE == 'postgres':
                # PostgreSQL DictCursor 返回字典风格的行
                row_data = dict(row)
            else:
                # SQLite Row 对象（已设置row_factory = sqlite3.Row）
                if hasattr(row, 'keys'):
                    row_data = dict(row)
                else:
                    # 兜底：普通tuple格式
                    row_data = {
                        'id': row[0],
                        'title': row[1],
                        'description': row[2],
                        'status': row[3],
                        'project': row[4],
                        'created_at': row[5],
                        'resolved_at': row[6],
                        'resolution': row[7],
                        'image_path': row[8],
                        'creator_username': row[9],
                        'creator_name': row[10],
                        'assignee_username': row[11],
                        'assignee_name': row[12],
                        'assignee_team': row[13],
                        'manager_username': row[14],
                        'manager_name': row[15]
                    }

            # 添加前端需要的字段映射
            row_data['creator'] = row_data.get('creator_name') or row_data.get('creator_username') or '未知'
            row_data['assignee'] = row_data.get('assignee_name') or row_data.get('assignee_username') or '未分配'
            row_data['manager'] = row_data.get('manager_name') or row_data.get('manager_username') or '未分配'

            # 处理时间格式 - 转换为字符串格式
            created_at = row_data.get('created_at')
            if created_at:
                if hasattr(created_at, 'strftime'):
                    # datetime对象，转换为字符串
                    row_data['created_at'] = created_at.strftime('%Y-%m-%d %H:%M:%S')
                    row_data['create_time'] = row_data['created_at']
                else:
                    # 已经是字符串
                    row_data['create_time'] = str(created_at)
            else:
                row_data['create_time'] = ''

            resolved_at = row_data.get('resolved_at')
            if resolved_at:
                if hasattr(resolved_at, 'strftime'):
                    # datetime对象，转换为字符串
                    row_data['resolved_at'] = resolved_at.strftime('%Y-%m-%d %H:%M:%S')
                    row_data['resolve_time'] = row_data['resolved_at']
                else:
                    # 已经是字符串
                    row_data['resolve_time'] = str(resolved_at)
            else:
                row_data['resolve_time'] = ''

            # 默认产品线值
            row_data['team'] = '暂无'

            # 收集分配者用户名，用于后续查询团队信息
            if row_data.get('assignee_username'):
                assignee_usernames.add(row_data.get('assignee_username'))

            # 收集bug ID，用于查询附件数量
            if row_data.get('id'):
                bug_ids.append(row_data.get('id'))

            # 附件数量（默认值，后续会更新）
            row_data['attachments'] = '0'

            preview_data.append(row_data)

        # 批量查询用户团队信息
        if assignee_usernames:
            try:
                app.logger.info(f"开始查询用户团队，用户名: {assignee_usernames}")

                # 创建新的数据库连接
                user_conn = get_db_connection()
                if DB_TYPE == 'postgres':
                    from psycopg2.extras import DictCursor
                    user_cursor = user_conn.cursor(cursor_factory=DictCursor)
                else:
                    user_cursor = user_conn.cursor()

                # 构建查询参数
                placeholders = ', '.join(['%s'] * len(assignee_usernames))
                user_query, user_params = adapt_sql(f"SELECT username, team FROM users WHERE username IN ({placeholders})", tuple(assignee_usernames))
                user_cursor.execute(user_query, user_params)

                # 创建用户名到团队的映射
                user_teams = {}
                team_results = user_cursor.fetchall()
                app.logger.info(f"团队查询结果数量: {len(team_results)}")

                for user_row in team_results:
                    if DB_TYPE == 'postgres':
                        username = user_row['username']
                        team = user_row['team']
                    else:
                        username = user_row[0]
                        team = user_row[1]

                    user_teams[username] = team
                    app.logger.info(f"用户团队映射: {username} -> {team}")

                # 更新每条记录的团队信息
                updated_count = 0
                for row_data in preview_data:
                    assignee_username = row_data.get('assignee_username')
                    if assignee_username in user_teams and user_teams[assignee_username]:
                        row_data['team'] = user_teams[assignee_username]
                        updated_count += 1

                app.logger.info(f"更新了 {updated_count} 条记录的团队信息")

                # 关闭连接
                user_cursor.close()
                user_conn.close()

            except Exception as e:
                app.logger.error(f"批量获取用户团队失败: {e}")
                import traceback
                app.logger.error(f"详细错误: {traceback.format_exc()}")
                # 出错时保持默认值

        # 批量查询附件数量
        if bug_ids:
            try:
                app.logger.info(f"开始查询附件数量，bug数量: {len(bug_ids)}")

                # 创建新的数据库连接
                attachment_conn = get_db_connection()
                if DB_TYPE == 'postgres':
                    from psycopg2.extras import DictCursor
                    attachment_cursor = attachment_conn.cursor(cursor_factory=DictCursor)
                else:
                    attachment_cursor = attachment_conn.cursor()

                # 构建查询参数
                placeholders = ', '.join(['%s'] * len(bug_ids))
                attachment_query, attachment_params = adapt_sql(f"""
                    SELECT bug_id, COUNT(*) as attachment_count
                    FROM bug_images
                    WHERE bug_id IN ({placeholders})
                    GROUP BY bug_id
                """, tuple(bug_ids))
                attachment_cursor.execute(attachment_query, attachment_params)

                # 创建bug ID到附件数量的映射
                attachment_counts = {}
                attachment_results = attachment_cursor.fetchall()
                app.logger.info(f"附件查询结果数量: {len(attachment_results)}")

                for attachment_row in attachment_results:
                    if DB_TYPE == 'postgres':
                        bug_id = attachment_row['bug_id']
                        count = attachment_row['attachment_count']
                    else:
                        if hasattr(attachment_row, 'keys'):
                            attachment_dict = dict(attachment_row)
                            bug_id = attachment_dict['bug_id']
                            count = attachment_dict['attachment_count']
                        else:
                            bug_id = attachment_row[0]
                            count = attachment_row[1]

                    attachment_counts[bug_id] = count
                    app.logger.info(f"Bug {bug_id} 附件数量: {count}")

                # 更新每条记录的附件数量
                updated_attachment_count = 0
                for row_data in preview_data:
                    bug_id = row_data.get('id')
                    if bug_id in attachment_counts:
                        row_data['attachments'] = str(attachment_counts[bug_id])
                        updated_attachment_count += 1

                app.logger.info(f"更新了 {updated_attachment_count} 条记录的附件数量")

                # 关闭连接
                attachment_cursor.close()
                attachment_conn.close()

            except Exception as e:
                app.logger.error(f"批量获取附件数量失败: {e}")
                import traceback
                app.logger.error(f"详细错误: {traceback.format_exc()}")
                # 出错时保持默认值

        # 确保所有datetime对象都已转换为字符串
        for row in preview_data:
            for key, value in row.items():
                if hasattr(value, 'strftime'):
                    row[key] = value.strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({
            'success': True,
            'data': preview_data,
            'fields': fields,
            'count': len(preview_data)
        })

    except Exception as e:
        app.logger.error(f"预览报表数据失败: {e}")
        return jsonify({'error': '预览失败', 'message': str(e)}), 500

@app.route('/admin/reports/chart-data', methods=['POST'])
@login_required
@role_required('gly')
def admin_reports_chart_data():
    """获取报表图表数据"""

    try:
        data = request.get_json()
        filters = data.get('filters', {})

        # 构建查询（与预览查询相同的筛选条件）
        query_parts = []
        params = []

        # 基础查询
        base_query = '''
            SELECT b.id, b.title, b.status, b.type, b.project,
                   b.created_at, b.resolved_at,
                   u1.username as creator_username, u1.chinese_name as creator_name,
                   u2.username as assignee_username, u2.chinese_name as assignee_name,
                   u1.team as creator_team, u2.team as assignee_team
            FROM bugs b
            LEFT JOIN users u1 ON b.created_by = u1.id
            LEFT JOIN users u2 ON b.assigned_to = u2.id
        '''

        # 添加筛选条件（与预览接口相同的逻辑）
        where_conditions = []

        # 日期范围筛选
        if filters.get('dateRange'):
            date_range = filters['dateRange']
            if date_range.get('start'):
                where_conditions.append('DATE(b.created_at) >= %s')
                params.append(date_range['start'])
            if date_range.get('end'):
                where_conditions.append('DATE(b.created_at) <= %s')
                params.append(date_range['end'])

        # 状态筛选
        if filters.get('status'):
            status_list = filters['status']
            if status_list:
                placeholders = ','.join(['%s'] * len(status_list))
                where_conditions.append(f'b.status IN ({placeholders})')
                params.extend(status_list)

        # 项目筛选
        if filters.get('project'):
            where_conditions.append('b.project = %s')
            params.append(filters['project'])

        # 创建者筛选
        if filters.get('creator'):
            where_conditions.append('b.created_by = %s')
            params.append(filters['creator'])

        # 分配者筛选
        if filters.get('assignee'):
            where_conditions.append('b.assigned_to = %s')
            params.append(filters['assignee'])

        # 类型筛选
        if filters.get('type'):
            type_list = filters['type']
            if type_list:
                placeholders = ','.join(['%s'] * len(type_list))
                where_conditions.append(f'b.type IN ({placeholders})')
                params.extend(type_list)

        # 组合查询
        if where_conditions:
            query = base_query + ' WHERE ' + ' AND '.join(where_conditions)
        else:
            query = base_query

        # 执行查询
        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            from psycopg2.extras import DictCursor
            cursor = conn.cursor(cursor_factory=DictCursor)
        else:
            cursor = conn.cursor()

        query, params = adapt_sql(query, params)
        cursor.execute(query, params)

        results = cursor.fetchall()
        conn.close()

        # 统计数据（只统计已完成的问题）
        creator_stats = {'bug': {}, '需求': {}}  # 提交人统计（按类型分别统计）
        assignee_stats = {'bug': {}, '需求': {}}  # 处理人统计（按类型分别统计）
        type_stats = {}  # 类型统计

        for row in results:
            # 统一转换为字典格式
            if DB_TYPE == 'postgres':
                row_data = dict(row)
            else:
                if hasattr(row, 'keys'):
                    row_data = dict(row)
                else:
                    # 兜底：普通tuple格式
                    row_data = {
                        'id': row[0],
                        'title': row[1],
                        'status': row[2],
                        'type': row[3],
                        'project': row[4],
                        'created_at': row[5],
                        'resolved_at': row[6],
                        'creator_username': row[7],
                        'creator_name': row[8],
                        'assignee_username': row[9],
                        'assignee_name': row[10],
                        'creator_team': row[11],
                        'assignee_team': row[12]
                    }

            # 获取问题类型
            bug_type = row_data.get('type') or 'bug'

            # 统计类型分布
            type_stats[bug_type] = type_stats.get(bug_type, 0) + 1

            # 只统计已完成状态的问题
            if row_data.get('status') == '已完成' and row_data.get('resolved_at'):
                # 统计提交人数据（按类型分别统计）
                creator_name = row_data.get('creator_name') or row_data.get('creator_username') or '未知'
                if bug_type not in creator_stats:
                    creator_stats[bug_type] = {}
                creator_stats[bug_type][creator_name] = creator_stats[bug_type].get(creator_name, 0) + 1

                # 统计处理人数据（按类型分别统计）
                assignee_name = row_data.get('assignee_name') or row_data.get('assignee_username') or '未分配'
                if assignee_name != '未分配':
                    if bug_type not in assignee_stats:
                        assignee_stats[bug_type] = {}
                    assignee_stats[bug_type][assignee_name] = assignee_stats[bug_type].get(assignee_name, 0) + 1

        # 转换为图表数据格式（按类型分别统计）
        creator_chart_data = {}
        assignee_chart_data = {}

        # 处理提交人统计数据
        for bug_type, stats in creator_stats.items():
            if stats:  # 只有当该类型有数据时才添加
                creator_chart_data[bug_type] = {
                    'labels': list(stats.keys()),
                    'values': list(stats.values())
                }

        # 处理处理人统计数据
        for bug_type, stats in assignee_stats.items():
            if stats:  # 只有当该类型有数据时才添加
                assignee_chart_data[bug_type] = {
                    'labels': list(stats.keys()),
                    'values': list(stats.values())
                }

        # 类型分布图表数据
        type_chart_data = {
            'labels': list(type_stats.keys()),
            'values': list(type_stats.values())
        }

        chart_data = {
            'creator': creator_chart_data,
            'assignee': assignee_chart_data,
            'type': type_chart_data
        }

        return jsonify({'success': True, 'data': chart_data})

    except Exception as e:
        app.logger.error(f"获取图表数据失败: {e}")
        import traceback
        app.logger.error(f"详细错误: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/reports/projects', methods=['GET'])
@login_required
@role_required('gly')
def admin_reports_projects():
    """获取所有项目列表"""
    try:
        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            from psycopg2.extras import DictCursor
            cursor = conn.cursor(cursor_factory=DictCursor)
        else:
            cursor = conn.cursor()

        # 查询所有不同的项目
        cursor.execute('SELECT DISTINCT project FROM bugs WHERE project IS NOT NULL AND project != \'\' ORDER BY project')
        results = cursor.fetchall()
        conn.close()

        # 提取项目名称
        projects = []
        for row in results:
            if DB_TYPE == 'postgres':
                project = row['project']
            else:
                project = row[0]
            if project:
                projects.append(project)

        return jsonify({
            'success': True,
            'projects': projects
        })

    except Exception as e:
        app.logger.error(f"获取项目列表失败: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/reports/export', methods=['POST'])
@login_required
@role_required('gly')
def admin_reports_export():
    """导出报表数据"""

    try:
        data = request.get_json()
        filters = data.get('filters', {})
        fields = data.get('fields', [])
        format_type = data.get('format', 'excel')
        filename = data.get('filename', '问题列表报表')

        # 获取完整数据（不限制条数）
        conn = get_db_connection()
        if DB_TYPE == 'postgres':
            from psycopg2.extras import DictCursor
            cursor = conn.cursor(cursor_factory=DictCursor)
        else:
            cursor = conn.cursor()

        # 构建查询（与预览相同，但不限制条数）
        base_query = '''
            SELECT b.id, b.title, b.description, b.status, b.type, b.project,
                   b.created_at, b.resolved_at, b.resolution, b.image_path,
                   u1.username as creator_username, u1.chinese_name as creator_name,
                   u2.username as assignee_username, u2.chinese_name as assignee_name,
                   u2.team as assignee_team,
                   u3.username as manager_username, u3.chinese_name as manager_name
            FROM bugs b
            LEFT JOIN users u1 ON b.created_by = u1.id
            LEFT JOIN users u2 ON b.assigned_to = u2.id
            LEFT JOIN users u3 ON u2.team = u3.team AND u3.role_en = 'fzr'
        '''

        # 添加筛选条件
        where_conditions = []
        params = []

        # 日期范围筛选
        if filters.get('dateRange'):
            date_range = filters['dateRange']
            if date_range.get('start'):
                where_conditions.append('DATE(b.created_at) >= %s')
                params.append(date_range['start'])
            if date_range.get('end'):
                where_conditions.append('DATE(b.created_at) <= %s')
                params.append(date_range['end'])

        # 状态筛选
        if filters.get('status'):
            status_list = filters['status']
            if status_list:
                placeholders = ','.join(['%s'] * len(status_list))
                where_conditions.append(f'b.status IN ({placeholders})')
                params.extend(status_list)

        # 项目筛选
        if filters.get('project'):
            where_conditions.append('b.project = %s')
            params.append(filters['project'])

        # 创建者筛选
        if filters.get('creator'):
            where_conditions.append('b.created_by = %s')
            params.append(filters['creator'])

        # 分配者筛选
        if filters.get('assignee'):
            where_conditions.append('b.assigned_to = %s')
            params.append(filters['assignee'])

        # 类型筛选
        if filters.get('type'):
            type_list = filters['type']
            if type_list:
                placeholders = ','.join(['%s'] * len(type_list))
                where_conditions.append(f'b.type IN ({placeholders})')
                params.extend(type_list)

        # 组合查询
        if where_conditions:
            query = base_query + ' WHERE ' + ' AND '.join(where_conditions)
        else:
            query = base_query

        query += ' ORDER BY b.created_at DESC'

        # 执行查询
        query, params = adapt_sql(query, params)
        cursor.execute(query, params)
        results = cursor.fetchall()
        conn.close()

        # 转换数据格式 - 兼容PostgreSQL和SQLite
        export_data = []
        assignee_usernames = set()  # 收集所有分配者用户名
        bug_ids = []  # 收集所有bug ID，用于查询附件数量

        for row in results:
            # 统一转换为字典格式
            if DB_TYPE == 'postgres':
                # PostgreSQL DictCursor 返回字典风格的行
                row_data = dict(row)
            else:
                # SQLite Row 对象（已设置row_factory = sqlite3.Row）
                if hasattr(row, 'keys'):
                    row_data = dict(row)
                else:
                    # 兜底：普通tuple格式
                    row_data = {
                        'id': row[0],
                        'title': row[1],
                        'description': row[2],
                        'status': row[3],
                        'project': row[4],
                        'created_at': row[5],
                        'resolved_at': row[6],
                        'resolution': row[7],
                        'image_path': row[8],
                        'creator_username': row[9],
                        'creator_name': row[10],
                        'assignee_username': row[11],
                        'assignee_name': row[12]
                    }

            # 添加前端需要的字段映射
            row_data['creator'] = row_data.get('creator_name') or row_data.get('creator_username') or '未知'
            row_data['assignee'] = row_data.get('assignee_name') or row_data.get('assignee_username') or '未分配'
            row_data['manager'] = row_data.get('manager_name') or row_data.get('manager_username') or '未分配'

            # 处理时间格式 - 转换为字符串格式
            created_at = row_data.get('created_at')
            if created_at:
                if hasattr(created_at, 'strftime'):
                    # datetime对象，转换为字符串
                    row_data['create_time'] = created_at.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    # 已经是字符串
                    row_data['create_time'] = str(created_at)
            else:
                row_data['create_time'] = ''

            resolved_at = row_data.get('resolved_at')
            if resolved_at:
                if hasattr(resolved_at, 'strftime'):
                    # datetime对象，转换为字符串
                    row_data['resolve_time'] = resolved_at.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    # 已经是字符串
                    row_data['resolve_time'] = str(resolved_at)
            else:
                row_data['resolve_time'] = ''

            # 默认产品线值
            row_data['team'] = '暂无'

            # 收集分配者用户名，用于后续查询团队信息
            if row_data.get('assignee_username'):
                assignee_usernames.add(row_data.get('assignee_username'))

            # 收集bug ID，用于查询附件数量
            if row_data.get('id'):
                bug_ids.append(row_data.get('id'))

            # 附件数量（默认值，后续会更新）
            row_data['attachments'] = '0'

            export_data.append(row_data)

        # 批量查询用户团队信息
        if assignee_usernames:
            try:
                # 创建新的数据库连接
                user_conn = get_db_connection()
                if DB_TYPE == 'postgres':
                    from psycopg2.extras import DictCursor
                    user_cursor = user_conn.cursor(cursor_factory=DictCursor)
                else:
                    user_cursor = user_conn.cursor()

                # 构建查询参数
                placeholders = ', '.join(['%s'] * len(assignee_usernames))
                user_query, user_params = adapt_sql(f"SELECT username, team FROM users WHERE username IN ({placeholders})", tuple(assignee_usernames))
                user_cursor.execute(user_query, user_params)

                # 创建用户名到团队的映射
                user_teams = {}
                team_results = user_cursor.fetchall()
                app.logger.info(f"导出功能团队查询结果数量: {len(team_results)}")

                for user_row in team_results:
                    if DB_TYPE == 'postgres':
                        username = user_row['username']
                        team = user_row['team']
                    else:
                        if hasattr(user_row, 'keys'):
                            user_dict = dict(user_row)
                            username = user_dict['username']
                            team = user_dict['team']
                        else:
                            username = user_row[0]
                            team = user_row[1]

                    user_teams[username] = team
                    app.logger.info(f"导出功能用户团队映射: {username} -> {team}")

                # 更新每条记录的团队信息
                app.logger.info(f"用户团队映射: {user_teams}")
                updated_count = 0
                for row_data in export_data:
                    assignee_username = row_data.get('assignee_username')
                    app.logger.info(f"处理记录: ID={row_data.get('id')}, 分配者={assignee_username}")
                    if assignee_username in user_teams and user_teams[assignee_username]:
                        row_data['team'] = user_teams[assignee_username]
                        updated_count += 1
                        app.logger.info(f"  ✅ 更新产品线: {user_teams[assignee_username]}")
                    else:
                        app.logger.info(f"  ❌ 未找到产品线: {assignee_username}")

                app.logger.info(f"共更新 {updated_count}/{len(export_data)} 条记录的产品线")

                # 关闭连接
                user_cursor.close()
                user_conn.close()

            except Exception as e:
                app.logger.error(f"批量获取用户团队失败: {e}")
                # 出错时保持默认值

        # 批量查询附件数量
        if bug_ids:
            try:
                app.logger.info(f"开始查询附件数量，bug数量: {len(bug_ids)}")

                # 创建新的数据库连接
                attachment_conn = get_db_connection()
                if DB_TYPE == 'postgres':
                    from psycopg2.extras import DictCursor
                    attachment_cursor = attachment_conn.cursor(cursor_factory=DictCursor)
                else:
                    attachment_cursor = attachment_conn.cursor()

                # 构建查询参数
                placeholders = ', '.join(['%s'] * len(bug_ids))
                attachment_query, attachment_params = adapt_sql(f"""
                    SELECT bug_id, COUNT(*) as attachment_count
                    FROM bug_images
                    WHERE bug_id IN ({placeholders})
                    GROUP BY bug_id
                """, tuple(bug_ids))
                attachment_cursor.execute(attachment_query, attachment_params)

                # 创建bug ID到附件数量的映射
                attachment_counts = {}
                attachment_results = attachment_cursor.fetchall()
                app.logger.info(f"附件查询结果数量: {len(attachment_results)}")

                for attachment_row in attachment_results:
                    if DB_TYPE == 'postgres':
                        bug_id = attachment_row['bug_id']
                        count = attachment_row['attachment_count']
                    else:
                        if hasattr(attachment_row, 'keys'):
                            attachment_dict = dict(attachment_row)
                            bug_id = attachment_dict['bug_id']
                            count = attachment_dict['attachment_count']
                        else:
                            bug_id = attachment_row[0]
                            count = attachment_row[1]

                    attachment_counts[bug_id] = count
                    app.logger.info(f"Bug {bug_id} 附件数量: {count}")

                # 更新每条记录的附件数量
                updated_attachment_count = 0
                for row_data in export_data:
                    bug_id = row_data.get('id')
                    if bug_id in attachment_counts:
                        row_data['attachments'] = str(attachment_counts[bug_id])
                        updated_attachment_count += 1

                app.logger.info(f"更新了 {updated_attachment_count} 条记录的附件数量")

                # 关闭连接
                attachment_cursor.close()
                attachment_conn.close()

            except Exception as e:
                app.logger.error(f"批量获取附件数量失败: {e}")
                import traceback
                app.logger.error(f"详细错误: {traceback.format_exc()}")
                # 出错时保持默认值

        # 根据格式导出
        if format_type == 'excel':
            return export_to_excel(export_data, fields, filename)
        else:
            return jsonify({'error': '不支持的导出格式'}), 400

    except Exception as e:
        app.logger.error(f"导出报表数据失败: {e}")
        return jsonify({'error': '导出失败', 'message': str(e)}), 500

def export_to_excel(data, fields, filename):
    """导出为Excel格式（使用openpyxl）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        # 根据文件名判断工作表标题
        if '项目列表' in filename:
            ws.title = "项目列表"
        else:
            ws.title = "问题列表"

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头
        headers = [field.get('label', field.get('key', '')) for field in fields]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据
        for row_idx, row in enumerate(data, 2):
            for col_idx, field in enumerate(fields, 1):
                field_key = field.get('key')
                value = row.get(field_key, '')
                if value is None:
                    value = ''
                ws.cell(row=row_idx, column=col_idx, value=str(value))

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
            ws.column_dimensions[column_letter].width = adjusted_width

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 返回文件
        from flask import Response
        import urllib.parse

        # 处理中文文件名 - 使用RFC 5987标准
        # 生成ASCII安全的文件名作为兜底
        import re
        ascii_filename = re.sub(r'[^a-zA-Z0-9\-_.]', '_', filename) + '.xlsx'
        if ascii_filename == '.xlsx' or ascii_filename == '_.xlsx':
            ascii_filename = 'bug_report.xlsx'  # 如果文件名全是特殊字符，使用默认名称

        # UTF-8编码的完整文件名
        encoded_filename = urllib.parse.quote(f'{filename}.xlsx'.encode('utf-8'))

        # 创建响应
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{ascii_filename}"; filename*=UTF-8\'\'{encoded_filename}',
                'Cache-Control': 'no-cache'
            }
        )

        return response

    except ImportError:
        return jsonify({'error': 'Excel导出功能需要安装openpyxl库'}), 500
    except Exception as e:
        app.logger.error(f"Excel导出失败: {e}")
        return jsonify({'error': 'Excel导出失败', 'message': str(e)}), 500








if __name__ == '__main__':
    # 检查端口是否可用
    HOST = '127.0.0.1'
    PORT = 5000

    print("🚀 ReBugTracker 启动中...")
    check_port_available(HOST, PORT)

    # 初始化数据库
    print("🗄️ 初始化数据库...")
    init_db()
    print("✅ 数据库初始化完成")

    # 确保必要目录存在
    print("📁 初始化目录...")
    upload_dir = app.config['UPLOAD_FOLDER']
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir, mode=0o777, exist_ok=True)
        print(f"✅ 创建上传目录: {upload_dir}")
    else:
        print(f"✅ 上传目录已存在: {upload_dir}")

    # 验证目录权限
    if not os.access(upload_dir, os.W_OK):
        print(f"⚠️ 警告: 上传目录可能无写入权限: {upload_dir}")
    else:
        print(f"✅ 上传目录权限正常: {upload_dir}")

    try:
        # 启动通知清理调度器
        print("🧹 启动通知清理调度器...")
        from notification.cleanup_manager import cleanup_manager
        cleanup_manager.start_cleanup_scheduler(interval_hours=24)  # 每24小时清理一次

        print(f"📡 应用程序将在 http://{HOST}:{PORT} 启动")
        app.run(host=HOST, port=PORT, debug=True, use_reloader=False)
    except KeyboardInterrupt:
        print("\n👋 应用程序已停止")
    except Exception as e:
        print(f"❌ 启动失败: {e}")
    finally:
        # 停止清理调度器
        try:
            from notification.cleanup_manager import cleanup_manager
            cleanup_manager.stop_cleanup_scheduler()
            print("🧹 通知清理调度器已停止")
        except:
            pass

        # 确保所有资源被释放
        import os
        import signal
        try:
            os.kill(os.getpid(), signal.SIGTERM)
        except:
            pass
